"""Ветка без исходной сметы: предложение, согласие и готовый XLSX."""

import json
from pathlib import Path

from openpyxl import load_workbook

from _mvp_fixtures import (
    ConstructionAuditMVP,
    analysis,
    call,
    ctx,
    descriptor,
    make_plan,
    price_catalog_response,
    mvp_api,
)


def _setup_plan_only(mvp_api, tmp_path, *, confirm=False, catalog=False):
    ext = ConstructionAuditMVP(mvp_api)
    assert call(ext, "create_case", job_id="estimate_1", object_name="Офис")["ok"]
    plan = make_plan(tmp_path)
    imported = call(ext, "import_plan", job_id="estimate_1", plan=descriptor(plan))
    assert imported["ok"], imported
    saved = call(
        ext,
        "save_geometry",
        ctx("plan-review", "plan-review-msg"),
        job_id="estimate_1",
        analysis=analysis(),
    )
    assert saved["ok"], saved
    if confirm:
        confirmed = call(
            ext,
            "confirm_geometry",
            ctx("plan-confirm", "plan-confirm-msg"),
            job_id="estimate_1",
            geometry_revision=saved["geometry_revision"],
            confirmed=True,
        )
        assert confirmed["ok"], confirmed
    if catalog:
        offered = call(
            ext,
            "save_price_catalog",
            job_id="estimate_1",
            catalog_response=price_catalog_response(),
        )
        assert offered["ok"], offered
        return ext, imported, saved, offered
    return ext, imported, saved


def test_plan_only_import_does_not_create_estimate_artifacts(mvp_api, tmp_path):
    _, imported, _ = _setup_plan_only(mvp_api, tmp_path)
    root = mvp_api.root / "state/jobs/estimate_1"
    manifest = json.loads((root / "manifest.json").read_text())
    assert imported["status"] == "plan_imported"
    assert manifest["input_mode"] == "plan_only"
    assert set(manifest["documents"]) == {"plan"}
    assert not (root / "output/estimate_normalized.json").exists()
    assert not (root / "assets/estimate.xlsx").exists()


def test_plan_only_catalog_returns_offer_instead_of_mapping(mvp_api, tmp_path):
    _, _, _, offered = _setup_plan_only(mvp_api, tmp_path, confirm=True, catalog=True)
    assert offered["status"] == "estimate_generation_offer"
    assert offered["next_action"] == "await_estimate_generation_confirmation"
    assert "Сформировать смету" in offered["offer_markdown"]
    assert "mapping" not in offered


def test_generate_estimate_requires_confirmed_geometry_and_catalog(mvp_api, tmp_path):
    ext, _, saved = _setup_plan_only(mvp_api, tmp_path)
    blocked = call(ext, "generate_estimate", job_id="estimate_1")
    assert blocked["code"] == "geometry_confirmation_required"

    confirmed = call(
        ext,
        "confirm_geometry",
        ctx("plan-confirm", "plan-confirm-msg"),
        job_id="estimate_1",
        geometry_revision=saved["geometry_revision"],
        confirmed=True,
    )
    assert confirmed["ok"]
    blocked = call(ext, "generate_estimate", job_id="estimate_1")
    assert blocked["code"] == "price_catalog_required"


def test_generated_estimate_matches_mvp_template_and_deduplicates_doors(mvp_api, tmp_path):
    ext, _, _, _ = _setup_plan_only(mvp_api, tmp_path, confirm=True, catalog=True)
    result = call(ext, "generate_estimate", job_id="estimate_1")
    assert result["status"] == "estimate_generated"
    assert result["rows_count"] == 13

    artifact = Path(result["estimate_artifact"]["path"])
    assert artifact.is_file()
    workbook = load_workbook(artifact, data_only=False)
    try:
        assert workbook.sheetnames == ["Смета"]
        sheet = workbook["Смета"]
        assert [cell.value for cell in sheet[1]] == [
            "№", "Помещение", "Наименование работы", "Единица", "Количество",
            "Цена за единицу", "Стоимость", "Примечание",
        ]
        assert sheet.row_dimensions[1].height == 42.75
        assert sheet.column_dimensions["H"].width == 34
        assert sheet.tables["GeneratedEstimateTable"].ref == "A1:H14"
        assert sheet.tables["GeneratedEstimateTable"].tableStyleInfo.name == "TableStyleMedium2"
        assert sheet["G2"].value == "=E2*F2"
        assert "MCP work id:" in sheet["F2"].comment.text

        door_rows = [
            row for row in sheet.iter_rows(min_row=2, values_only=True)
            if row[2] == "Установка дверей"
        ]
        assert len(door_rows) == 1
        assert door_rows[0][1] == "Весь объект"
        assert door_rows[0][4] == 1
    finally:
        workbook.close()


def test_geometry_correction_invalidates_generated_estimate(mvp_api, tmp_path):
    ext, _, _, _ = _setup_plan_only(mvp_api, tmp_path, confirm=True, catalog=True)
    generated = call(ext, "generate_estimate", job_id="estimate_1")
    artifact = Path(generated["estimate_artifact"]["path"])
    assert artifact.is_file()

    corrected = call(
        ext,
        "confirm_geometry",
        ctx("plan-correction", "plan-correction-msg"),
        job_id="estimate_1",
        geometry_revision=1,
        confirmed=False,
        corrections=[{
            "target": "rooms",
            "room_ids": ["room_001"],
            "element_ids": "all",
            "field": "height_m",
            "value": 3.2,
        }],
        user_statement="Высота первого офиса 3,2 метра",
    )
    assert corrected["geometry_revision"] == 2
    assert not artifact.exists()
