from __future__ import annotations

import hashlib
import json
import math
import os
import re
import secrets
import shutil
import tempfile
import zipfile
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path, PurePosixPath
from typing import Any

from . import insights, report, vision, visual


SKILL_VERSION = "0.7.7"
JOB_ID_RE = re.compile(r"^[A-Za-z0-9_-]{1,64}$")
SUBAGENT_TASK_ID_RE = re.compile(r"^[0-9a-f]{8}$")
MAX_XLSX_SIZE = 10 * 1024 * 1024
MAX_IMAGE_SIZE = 20 * 1024 * 1024
MAX_ZIP_ENTRIES = 2000
MAX_ZIP_UNCOMPRESSED = 100 * 1024 * 1024
MAX_ZIP_ENTRY = 50 * 1024 * 1024
Q2 = Decimal("0.01")
REQUIRED_HEADERS = (
    "№",
    "Помещение",
    "Наименование работы",
    "Единица",
    "Количество",
    "Цена за единицу",
    "Стоимость",
    "Примечание",
)
SUPPORTED_WORKS: dict[str, tuple[str, tuple[str, ...], str]] = {
    "Грунтовка стен": ("net_wall_area_m2", ("м²", "м2"), "area"),
    "Окраска стен": ("net_wall_area_m2", ("м²", "м2"), "area"),
    "Устройство пола": ("floor_area_m2", ("м²", "м2"), "area"),
    "Отделка потолка": ("ceiling_area_m2", ("м²", "м2"), "area"),
    "Монтаж плинтуса": ("baseboard_length_m", ("м", "пог. м", "п.м"), "length"),
    "Установка дверей": ("door_count", ("шт", "шт."), "count"),
    "Установка окон": ("window_count", ("шт", "шт."), "count"),
}
DOOR_INSTALLATION_WORK = "Установка дверей"
DOOR_QUANTITY_SCOPE = "object_total_unique_doors"
ESTIMATE_DECLARED_ALLOCATION = "estimate_declared"
PRICE_CATALOG_TOOL = "mcp_construction_prices__get_supported_works"
PRICE_CATALOG_MAX_ITEMS = 5000
MAX_VISUAL_DELEGATION_RESPONSE_CHARS = 14_000
MAX_LLM_INSIGHTS_ATTEMPTS = 3
MAX_SUBAGENT_JSON_CHARS = 100_000
AUDIT_ARTIFACTS = (
    "mapping.json",
    "quantities.json",
    "calculation_trace.json",
    "price_checks.json",
    "findings.json",
    "visual_photos.json",
    "visual_photo_analyses.json",
    "visual_insights.json",
    "llm_context.json",
    "llm_insights.json",
    "report.html",
)
ESTIMATE_ARTIFACTS = ("generated_estimate.xlsx", "generated_estimate.json")
DERIVED_ARTIFACTS = ("price_catalog.json", *AUDIT_ARTIFACTS, *ESTIMATE_ARTIFACTS)
GEOMETRY_CORRECTION_FIELDS = {
    "rooms": set(vision.MEASUREMENT_LIMITS),
    "doors": {"width_m", "height_m"},
    "windows": {"width_m", "height_m"},
}
GEOMETRY_CORRECTION_MAX_ITEMS = 100
GEOMETRY_REVIEW_ASSISTANT_INSTRUCTION = (
    "Выведи review_markdown полностью. Если затем появится [SYSTEM REMINDER] с "
    "[SUBAGENT_RESULTS], предыдущий черновик НЕ виден пользователю: обязательно вызови tool с "
    "суффиксом render_geometry_review и выведи его review_markdown. Никогда не пиши, что review "
    "уже показан выше. Не подтверждай геометрию и не вызывай confirm_geometry в текущем task."
)
GEOMETRY_HANDOFF_BARRIER_DRAFT = "Геометрия сохранена. Подготавливаю проверочную таблицу."
GEOMETRY_HANDOFF_BARRIER_ASSISTANT_INSTRUCTION = (
    "Результат save_geometry намеренно не содержит review_markdown. Если после завершения "
    "Vision-субагента в текущем task ещё не было [SYSTEM REMINDER] с [SUBAGENT_RESULTS], "
    "не вызывай tools и ответь РОВНО: «Геометрия сохранена. Подготавливаю проверочную "
    "таблицу.» Runtime перехватит этот draft; пользователю он не виден. После reminder "
    "обязательно вызови tool с суффиксом render_geometry_review для текущей revision и выведи "
    "его review_markdown полностью. Если reminder этого Vision-субагента уже есть в контексте, "
    "вызови render_geometry_review сразу. Никогда не пиши, что review показан выше."
)
GEOMETRY_CONFIRMATION_NOTICE = (
    "Ожидается отдельное новое сообщение пользователя с подтверждением указанной геометрии."
)
GEOMETRY_CORRECTION_ASSISTANT_INSTRUCTION = (
    "Сразу вызови render_geometry_review с возвращённой geometry_revision, выведи "
    "review_markdown полностью и заверши task. Не подтверждай исправленную геометрию в "
    "текущем task: для подтверждения нужно отдельное новое сообщение пользователя."
)
GEOMETRY_REVIEW_TERMINAL_MARKER = "[ОЖИДАНИЕ НОВОГО СООБЩЕНИЯ ПОЛЬЗОВАТЕЛЯ]"
AUDIT_SUMMARY_ASSISTANT_INSTRUCTION = (
    "Выведи audit_summary_markdown полностью. После ЛЮБОГО [SYSTEM REMINDER] с "
    "[SUBAGENT_RESULTS], в том числе после finalize_audit или предыдущего render, считай прежний "
    "assistant draft невидимым: снова вызови tool с суффиксом render_audit_summary и выведи только "
    "его audit_summary_markdown. Никогда не отвечай о состоянии handoff и не пиши, что результаты "
    "уже показаны выше. Не открывай report без явной просьбы пользователя."
)
LLM_INSIGHTS_ASSISTANT_INSTRUCTION = (
    "llm_insights_delegation в этом ответе полный. Вызови schedule_subagent с ним строго 1:1, затем "
    "wait_task. Не ищи packet через list_files, search_code, read_file или shell, даже если после "
    "ответа tool виден transport truncation marker. Предпочтительно дословно передай в finalize_audit "
    "JSON-строку между внутренними маркерами без json.loads и повторной сборки; если выделение "
    "ненадёжно, передай полный неизменённый wait_task result. Добавь insights_task_id. Не "
    "создавай гипотезы основным агентом и не показывай raw JSON."
)
VISUAL_REVIEW_OFFER_MARKDOWN = (
    "Детерминированная проверка сметы подготовлена. Если есть фотографии объекта, "
    "приложите одним ZIP-архивом не более 5 файлов PNG/JPG/JPEG. По каждой фотографии "
    "будет запущен отдельный Vision-субагент. Если фотографий нет, напишите «без фото»."
)
VISUAL_REVIEW_OFFER_ASSISTANT_INSTRUCTION = (
    "Выведи visual_review_offer_markdown полностью и заверши task. Не запускай аналитического "
    "LLM-субагента в этом task. Дождись отдельного нового сообщения пользователя: ZIP с не более "
    "чем 5 фотографиями либо явного ответа «без фото»."
)
VISUAL_DELEGATION_ASSISTANT_INSTRUCTION = (
    "Запусти все packet из visual_delegations одним burst 1:1 и сохрани photo_id -> task_id. "
    "Затем последовательно: wait_task, дословная передача JSON-строки между "
    "внутренними маркерами без json.loads и повторной сборки (либо полного неизменённого wait_task "
    "result, если выделение ненадёжно), затем save_visual_analysis с теми же photo_id/task_id; только "
    "после успешного save переходи дальше. Не вызывай wait_tasks, recent_tasks, cancel_task, "
    "get_task_result или повторный import_site_photos. Основной агент фото не анализирует. Никогда не "
    "изменяй analysis и не отправляй отклонённый task_id повторно: при schema error запусти "
    "возвращённый visual_delegation строго 1:1 с новым task_id."
)
VISUAL_RETRY_ASSISTANT_INSTRUCTION = (
    "Не изменяй и не отправляй analysis повторно. Немедленно вызови schedule_subagent с "
    "details.visual_delegation строго 1:1, затем wait_task и save_visual_analysis с новым task_id."
)
ESTIMATE_OFFER_ASSISTANT_INSTRUCTION = (
    "Выведи offer_markdown полностью и заверши task. Не вызывай generate_estimate в этом task: "
    "дождись отдельного нового сообщения пользователя с явным согласием сформировать смету."
)


class AuditError(Exception):
    def __init__(self, code: str, message: str, details: dict[str, Any] | None = None):
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details or {}


def _decode_subagent_json_object(
    value: Any,
    *,
    code: str,
    label: str,
) -> dict[str, Any]:
    """Decode a child object without asking the parent model to rebuild it.

    Prefer the exact JSON substring emitted by the child. A structured object
    remains accepted for backwards compatibility. As a stability fallback,
    also accept the unmodified ``wait_task`` projection and deterministically
    select the marked FINAL ANSWER after the outcome block.
    """
    transport_details = {
        "allowed_next_action": "reextract_same_subagent_result",
        "next_action": "reextract_same_subagent_result",
        "assistant_instruction": (
            "Не запускай нового субагента и не изменяй его содержательный ответ. Заново передай "
            "точную JSON-строку между внутренними маркерами либо полный неизменённый результат того "
            "же wait_task и повтори текущий tool с тем же task ID. Если исходный результат больше "
            "недоступен, остановись "
            "с orchestration_packet_invalid."
        ),
    }
    if isinstance(value, dict):
        return value
    if not isinstance(value, str):
        raise AuditError(
            code,
            f"{label} должен быть JSON object или его точной JSON-строкой.",
            transport_details,
        )
    if not value.strip() or len(value) > MAX_SUBAGENT_JSON_CHARS:
        raise AuditError(
            code,
            f"JSON-строка {label} пуста или превышает допустимый transport-размер.",
            transport_details,
        )
    serialized = value.strip()
    try:
        decoded = json.loads(serialized)
    except (json.JSONDecodeError, TypeError, ValueError) as exc:
        begin_marker = "[BEGIN_SUBTASK_OUTPUT]"
        end_marker = "[END_SUBTASK_OUTPUT]"
        final_prefix = f"FINAL ANSWER: {begin_marker}"

        # task_contract inside SUBTASK_OUTCOME may contain the same example
        # markers. Ignore that entire block and unwrap only the host-owned
        # output envelope that follows it.
        outcome_end = serialized.rfind("[/SUBTASK_OUTCOME]")
        if outcome_end >= 0:
            outer_begin = serialized.find(
                begin_marker, outcome_end + len("[/SUBTASK_OUTCOME]")
            )
            trace_begin = serialized.find(
                "[SUBTASK_TRACE]", outer_begin + len(begin_marker)
            )
            search_end = trace_begin if trace_begin >= 0 else len(serialized)
            outer_end = serialized.rfind(
                end_marker, outer_begin + len(begin_marker), search_end
            )
            if outer_begin < 0 or outer_end < 0:
                raise AuditError(
                    code,
                    f"Runtime envelope {label} не содержит завершённый subtask output.",
                    transport_details,
                ) from exc
            candidate = serialized[
                outer_begin + len(begin_marker):outer_end
            ].strip()
        else:
            candidate = serialized

        final_index = candidate.rfind(final_prefix)
        if final_index < 0:
            raise AuditError(
                code,
                f"JSON-строка {label} не разбирается и не содержит точный FINAL ANSWER envelope.",
                transport_details,
            ) from exc
        payload_start = final_index + len(final_prefix)
        payload_end = candidate.find(end_marker, payload_start)
        if payload_end < 0 or candidate[payload_end + len(end_marker):].strip():
            raise AuditError(
                code,
                f"FINAL ANSWER envelope {label} повреждён либо содержит postamble.",
                transport_details,
            ) from exc
        payload_text = candidate[payload_start:payload_end].strip()
        try:
            decoded = json.loads(payload_text)
        except (json.JSONDecodeError, TypeError, ValueError) as marked_exc:
            raise AuditError(
                code,
                f"JSON внутри FINAL ANSWER envelope {label} не разбирается без исправлений.",
                transport_details,
            ) from marked_exc
    if not isinstance(decoded, dict):
        raise AuditError(
            code,
            f"JSON-строка {label} должна содержать ровно один object.",
            transport_details,
        )
    return decoded


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def validate_job_id(job_id: Any) -> str:
    if not isinstance(job_id, str) or not JOB_ID_RE.fullmatch(job_id):
        raise AuditError("invalid_job_id", "job_id должен соответствовать ^[A-Za-z0-9_-]{1,64}$.")
    return job_id


def job_dir(api: Any, job_id: str) -> Path:
    validate_job_id(job_id)
    root = Path(api.skill_job_dir(job_id)).resolve(strict=False)
    root.mkdir(parents=True, exist_ok=True)
    for child in ("assets", "output", "tmp"):
        (root / child).mkdir(parents=True, exist_ok=True)
    return root


def output_path(root: Path, name: str) -> Path:
    return root / "output" / name


def atomic_write_bytes(path: Path, data: bytes) -> None:
    # Manifest и артефакты читаются в следующих tool-вызовах. Через os.replace
    # они видят либо старую версию целиком, либо новую, но не оборванную запись.
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary: str | None = None
    try:
        with tempfile.NamedTemporaryFile(dir=path.parent, prefix=f".{path.name}.", delete=False) as handle:
            temporary = handle.name
            handle.write(data)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
        temporary = None
    finally:
        if temporary:
            Path(temporary).unlink(missing_ok=True)


def atomic_write_json(path: Path, value: Any) -> None:
    try:
        payload = json.dumps(value, ensure_ascii=False, indent=2, allow_nan=False).encode("utf-8") + b"\n"
    except (TypeError, ValueError) as exc:
        raise AuditError("invalid_json_state", "Внутреннее состояние содержит недопустимые JSON-значения.") from exc
    atomic_write_bytes(path, payload)


def read_json(path: Path, *, code: str = "case_not_found") -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise AuditError(code, "Кейс или обязательный артефакт не найден.", {"file": path.name}) from exc
    except (OSError, json.JSONDecodeError) as exc:
        raise AuditError("storage_error", "Не удалось прочитать данные кейса.", {"file": path.name}) from exc


def load_manifest(root: Path) -> dict[str, Any]:
    value = read_json(root / "manifest.json")
    if not isinstance(value, dict):
        raise AuditError("storage_error", "Manifest кейса повреждён.")
    return value


def save_manifest(root: Path, manifest: dict[str, Any]) -> None:
    manifest["updated_at"] = now_iso()
    atomic_write_json(root / "manifest.json", manifest)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_json_sha256(value: Any) -> str:
    """Return a stable version fingerprint for one JSON representation."""
    payload = json.dumps(
        value,
        ensure_ascii=False,
        allow_nan=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _normalize_json_numbers(value: Any) -> Any:
    if isinstance(value, bool) or value is None or isinstance(value, (str, int)):
        return value
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return int(value)
        return value
    if isinstance(value, list):
        return [_normalize_json_numbers(item) for item in value]
    if isinstance(value, dict):
        return {key: _normalize_json_numbers(item) for key, item in value.items()}
    return value


def canonical_json_text(value: Any) -> str:
    return json.dumps(
        _normalize_json_numbers(value),
        ensure_ascii=False,
        allow_nan=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def create_case(api: Any, job_id: str, object_name: Any) -> tuple[dict[str, Any], bool]:
    validate_job_id(job_id)
    if not isinstance(object_name, str) or not 1 <= len(object_name.strip()) <= 200:
        raise AuditError("invalid_object_name", "object_name должен содержать от 1 до 200 символов.")
    object_name = object_name.strip()
    root = job_dir(api, job_id)
    path = root / "manifest.json"
    if path.exists():
        manifest = load_manifest(root)
        if manifest.get("job_id") != job_id or manifest.get("object_name") != object_name:
            raise AuditError("case_conflict", "Кейс с таким job_id уже создан для другого объекта.")
        return manifest, False
    created = now_iso()
    manifest = {
        "schema_version": 1,
        "skill_version": SKILL_VERSION,
        "job_id": job_id,
        "object_name": object_name,
        "documents_imported": False,
        "input_mode": None,
        "vision_attempts": 0,
        "geometry_revision": 0,
        "geometry_confirmed": False,
        "audit_completed": False,
        "audit_status": "not_started",
        "report_generated": False,
        "created_at": created,
        "updated_at": created,
        "documents": {},
    }
    atomic_write_json(path, manifest)
    return manifest, True


def _unlink(root: Path, names: tuple[str, ...]) -> None:
    for name in names:
        output_path(root, name).unlink(missing_ok=True)


def _invalidate_audit(manifest: dict[str, Any], root: Path) -> None:
    # Эти артефакты образуют одну зависимую цепочку: старый mapping нельзя
    # оставить рядом с новым каталогом или geometry, даже если сам файл валиден.
    _unlink(root, DERIVED_ARTIFACTS)
    manifest["audit_completed"] = False
    manifest["report_generated"] = False
    manifest["audit_status"] = "not_started"
    for key in (
        "audit_completed_at",
        "mapping_task_id",
        "mapping_sha256",
        "mapping_schema_version",
        "mapping_validation_status",
        "mapping_generation_token",
        "price_catalog_sha256",
        "price_catalog_items_count",
        "price_catalog_source_tool",
        "price_catalog_validation_status",
        "price_catalog_saved_at",
        "generated_estimate_sha256",
        "generated_estimate_geometry_sha256",
        "generated_estimate_price_catalog_sha256",
        "generated_estimate_rows_count",
        "generated_estimate_created_at",
        "deterministic_audit_prepared_at",
        "llm_context_sha256",
        "llm_insights_task_id",
        "llm_insights_sha256",
        "llm_insights_status",
        "llm_insights_completed_at",
        "llm_insights_attempts",
        "visual_review_status",
        "visual_photos_count",
        "visual_rejected_task_ids",
        "visual_insights_sha256",
        "visual_review_completed_at",
    ):
        manifest.pop(key, None)


def _invalidate_geometry(manifest: dict[str, Any], root: Path) -> None:
    # Новый план отменяет не только распознавание, но и пользовательское
    # подтверждение: оно относится к конкретной geometry revision и её hash.
    _unlink(
        root,
        ("geometry.json", "geometry_review.json", "geometry_corrections.json", *DERIVED_ARTIFACTS),
    )
    manifest["vision_attempts"] = 0
    manifest["geometry_confirmed"] = False
    manifest["audit_completed"] = False
    manifest["report_generated"] = False
    manifest["audit_status"] = "not_started"
    for key in (
        "review_task_id",
        "review_client_message_id",
        "confirmation_task_id",
        "confirmation_client_message_id",
        "confirmed_geometry_sha256",
        "geometry_confirmed_revision",
        "geometry_confirmed_at",
        "vision_task_id",
        "geometry_sha256",
        "geometry_schema_version",
        "geometry_validation_status",
        "geometry_correction_count",
        "geometry_corrections_sha256",
        "last_geometry_correction_at",
        "audit_completed_at",
        "mapping_task_id",
        "mapping_sha256",
        "mapping_schema_version",
        "mapping_validation_status",
        "mapping_generation_token",
        "price_catalog_sha256",
        "price_catalog_items_count",
        "price_catalog_source_tool",
        "price_catalog_validation_status",
        "price_catalog_saved_at",
        "generated_estimate_sha256",
        "generated_estimate_geometry_sha256",
        "generated_estimate_price_catalog_sha256",
        "generated_estimate_rows_count",
        "generated_estimate_created_at",
        "deterministic_audit_prepared_at",
        "llm_context_sha256",
        "llm_insights_task_id",
        "llm_insights_sha256",
        "llm_insights_status",
        "llm_insights_completed_at",
        "llm_insights_attempts",
        "visual_review_status",
        "visual_photos_count",
        "visual_rejected_task_ids",
        "visual_insights_sha256",
        "visual_review_completed_at",
    ):
        manifest.pop(key, None)


def _validate_attachment_descriptor(value: Any, expected_suffixes: set[str]) -> tuple[Path, str, str]:
    # Полям runtime не доверяем по отдельности: path и relpath должны описывать
    # один файл. Это отсекает подмену имени и простые traversal/symlink-сценарии.
    if not isinstance(value, dict) or set(value) != {"attachment_root", "attachment_relpath", "attachment_path"}:
        raise AuditError("invalid_attachment", "Attachment manifest entry должен содержать root, relpath и path одной записи.")
    root = value.get("attachment_root")
    relpath = value.get("attachment_relpath")
    absolute = value.get("attachment_path")
    if root != "artifact_store":
        raise AuditError("invalid_attachment", "attachment_root должен быть artifact_store.")
    if not isinstance(relpath, str) or "\\" in relpath:
        raise AuditError("invalid_attachment", "Некорректный attachment_relpath.")
    parts = relpath.split("/")
    pure = PurePosixPath(relpath)
    if pure.is_absolute() or len(parts) != 2 or parts[0] != "attachments" or any(part in {"", ".", ".."} for part in parts):
        raise AuditError("invalid_attachment", "Ожидается attachment_relpath вида attachments/<filename>.")
    filename = parts[1]
    suffix = Path(filename).suffix.lower()
    if filename.startswith(".") or suffix not in expected_suffixes:
        raise AuditError("invalid_attachment", "Расширение staged attachment не поддерживается.")
    if not isinstance(absolute, str) or not Path(absolute).is_absolute():
        raise AuditError("invalid_attachment", "attachment_path должен быть абсолютным staged path.")
    source = Path(absolute)
    if source.is_symlink():
        raise AuditError("invalid_attachment", "Символические ссылки не поддерживаются.")
    try:
        resolved = source.resolve(strict=True)
    except (OSError, RuntimeError) as exc:
        raise AuditError("attachment_not_found", "Staged attachment не найден.") from exc
    if not resolved.is_file() or resolved.parent.name != "attachments" or resolved.name != filename:
        raise AuditError("invalid_attachment", "Путь не соответствует указанной staged manifest entry.")
    return resolved, filename, suffix


def _validate_xlsx(path: Path) -> None:
    if path.stat().st_size > MAX_XLSX_SIZE:
        raise AuditError("attachment_too_large", "Размер XLSX превышает 10 MiB.")
    try:
        with zipfile.ZipFile(path, "r") as archive:
            # XLSX — это ZIP. Лимиты проверяем до openpyxl, иначе небольшой файл
            # может распаковаться в сотни мегабайт ещё до чтения таблицы.
            entries = archive.infolist()
            if len(entries) > MAX_ZIP_ENTRIES:
                raise AuditError("unsafe_xlsx_archive", "XLSX содержит слишком много ZIP entries.")
            total = 0
            names: set[str] = set()
            for entry in entries:
                name = entry.filename.replace("\\", "/")
                names.add(name)
                total += entry.file_size
                parts = PurePosixPath(name).parts
                if (
                    entry.flag_bits & 0x1
                    or entry.file_size > MAX_ZIP_ENTRY
                    or total > MAX_ZIP_UNCOMPRESSED
                    or PurePosixPath(name).is_absolute()
                    or any(part == ".." for part in parts)
                    or re.match(r"^[A-Za-z]:/", name)
                ):
                    raise AuditError("unsafe_xlsx_archive", "XLSX содержит небезопасный ZIP entry.", {"entry": entry.filename})
            if "[Content_Types].xml" not in names or not any(name.startswith("xl/") for name in names):
                raise AuditError("invalid_xlsx", "Файл не содержит структуру XLSX.")
            bad = archive.testzip()
            if bad:
                raise AuditError("invalid_xlsx", "ZIP-контейнер XLSX повреждён.", {"entry": bad})
    except (zipfile.BadZipFile, OSError) as exc:
        raise AuditError("invalid_xlsx", "Файл не является корректным XLSX.") from exc


def _image_mime(path: Path, suffix: str) -> str:
    if path.stat().st_size > MAX_IMAGE_SIZE:
        raise AuditError("attachment_too_large", "Размер изображения превышает 20 MiB.")
    head = path.read_bytes()[:16]
    if suffix == ".png" and head.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image/png"
    if suffix in {".jpg", ".jpeg"} and head.startswith(b"\xff\xd8\xff"):
        return "image/jpeg"
    raise AuditError("invalid_image", "Сигнатура изображения не соответствует PNG/JPG/JPEG.")


def _atomic_verified_copy(source: Path, destination: Path, expected_hash: str) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary: str | None = None
    try:
        with tempfile.NamedTemporaryFile(dir=destination.parent, prefix=f".{destination.name}.", delete=False) as target:
            temporary = target.name
            with source.open("rb") as source_handle:
                shutil.copyfileobj(source_handle, target)
            target.flush()
            os.fsync(target.fileno())
        # Staged-файл мог измениться между первичным хешированием и копированием.
        # В таком случае не публикуем копию под hash другого содержимого.
        if sha256_file(Path(temporary)) != expected_hash:
            raise AuditError("copy_verification_failed", "Не удалось подтвердить целостность копии документа.")
        os.replace(temporary, destination)
        temporary = None
    finally:
        if temporary:
            Path(temporary).unlink(missing_ok=True)


def _decimal(value: Any) -> Decimal | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, bool):
        return None
    try:
        if isinstance(value, float) and not math.isfinite(value):
            return None
        cleaned = str(value).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
        number = Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None
    return number if number.is_finite() else None


def decimal_text(value: Decimal | None) -> str | None:
    if value is None:
        return None
    if value == 0:
        return "0"
    return format(value.normalize(), "f")


def display_decimal(value: Decimal | str | None, *, integer: bool = False) -> str | None:
    if value is None:
        return None
    number = value if isinstance(value, Decimal) else Decimal(value)
    if integer and number == number.to_integral_value():
        return str(int(number))
    return format(number.quantize(Q2, rounding=ROUND_HALF_UP), "f")


def _service_row(values: dict[str, Any]) -> bool:
    work = str(values.get("Наименование работы") or "").strip().casefold().replace("ё", "е")
    room = str(values.get("Помещение") or "").strip().casefold().replace("ё", "е")
    joined = f"{room} {work}"
    markers = ("итого", "всего", "резерв", "примечан", "служебн", "заголов")
    if any(marker in joined for marker in markers):
        return True
    return work == "наименование работы" or room == "помещение"


def parse_estimate(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise AuditError("dependency_missing", "Для XLSX требуется openpyxl.") from exc
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except Exception as exc:
        raise AuditError("invalid_xlsx", "Не удалось безопасно открыть XLSX.") from exc
    try:
        if "Смета" not in workbook.sheetnames:
            raise AuditError("estimate_sheet_missing", "В XLSX отсутствует лист «Смета».")
        sheet = workbook["Смета"]
        header_row: int | None = None
        header_indexes: dict[str, int] = {}
        for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
            normalized = [str(value).strip() if value is not None else "" for value in row]
            if all(name in normalized for name in REQUIRED_HEADERS):
                if any(normalized.count(name) != 1 for name in REQUIRED_HEADERS):
                    raise AuditError("duplicate_header", "Обязательные заголовки должны встречаться ровно один раз.")
                header_row = row_index
                header_indexes = {name: normalized.index(name) for name in REQUIRED_HEADERS}
                break
        if header_row is None:
            raise AuditError("estimate_header_missing", "Не найдены обязательные заголовки листа «Смета».")

        rows: list[dict[str, Any]] = []
        warnings: list[dict[str, Any]] = []
        for source_row, raw in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            values = {name: raw[index] if index < len(raw) else None for name, index in header_indexes.items()}
            if all(value is None or (isinstance(value, str) and not value.strip()) for value in values.values()):
                continue
            if _service_row(values):
                warnings.append({"code": "service_row_skipped", "level": "info", "source_row": source_row, "message": "Служебная строка исключена из сметы для проверки."})
                continue
            position = values["№"]
            if position is None or not str(position).strip():
                warnings.append({"code": "service_row_skipped", "level": "info", "source_row": source_row, "message": "Строка без номера исключена из сметы для проверки."})
                continue
            room = str(values["Помещение"] or "").strip()
            work = str(values["Наименование работы"] or "").strip()
            unit = str(values["Единица"] or "").strip()
            issues: list[dict[str, str]] = []
            quantity = _decimal(values["Количество"])
            price = _decimal(values["Цена за единицу"])
            total = _decimal(values["Стоимость"])
            total_source = "provided_or_cached" if values["Стоимость"] not in (None, "") else "missing"
            if quantity is None or quantity < 0:
                issues.append({"type": "invalid_quantity", "field": "Количество"})
                quantity = None
            if values["Цена за единицу"] not in (None, "") and (price is None or price < 0):
                issues.append({"type": "invalid_price", "field": "Цена за единицу"})
                price = None
            if values["Стоимость"] not in (None, "") and (total is None or total < 0):
                issues.append({"type": "invalid_total", "field": "Стоимость"})
                total = None
            calculated_total = quantity * price if quantity is not None and price is not None else None
            # Заполненную стоимость не переписываем расчётной: расхождение в исходной
            # строке само является результатом проверки. Подставляем только пустую ячейку.
            if total is None and calculated_total is not None and not any(i["type"] == "invalid_total" for i in issues):
                total = calculated_total
                total_source = "calculated_from_quantity_and_price"
            elif total is not None and calculated_total is not None and abs(total - calculated_total) > Q2:
                issues.append({"type": "arithmetic_mismatch", "field": "Стоимость"})
            if not room:
                issues.append({"type": "unmapped_room", "field": "Помещение"})
            if not work:
                issues.append({"type": "unmapped_work", "field": "Наименование работы"})
            rows.append(
                {
                    "source_row": source_row,
                    "position": str(position).strip(),
                    "room": room,
                    "work_name": work,
                    "unit": unit,
                    "quantity": decimal_text(quantity),
                    "price": decimal_text(price),
                    "total": decimal_text(total),
                    "total_source": total_source,
                    "note": str(values["Примечание"] or "").strip(),
                    "issues": issues,
                }
            )
        if not rows:
            raise AuditError("estimate_empty", "В смете не найдено строк работ.")
        unique_rooms = sorted({row["room"] for row in rows if row["room"]}, key=str.casefold)
        unique_works = sorted({row["work_name"] for row in rows if row["work_name"]}, key=str.casefold)
        return {
            "schema_version": 1,
            "sheet": "Смета",
            "header_row": header_row,
            "rows": rows,
            "rooms": unique_rooms,
            "works": unique_works,
            "warnings": warnings,
        }
    finally:
        workbook.close()


def _vision_projection(api: Any, source: Path, suffix: str, source_hash: str) -> Path:
    info: dict[str, Any] = {}
    getter = getattr(api, "get_runtime_info", None)
    if callable(getter):
        try:
            candidate = getter()
            info = candidate if isinstance(candidate, dict) else {}
        except Exception:
            info = {}
    data_dir = str(info.get("data_dir") or "").strip()
    if not data_dir:
        return source
    # Субагенту нужен абсолютный путь в разрешённом uploads. Имя по hash не даёт
    # повторному запуску случайно открыть projection от другого плана.
    uploads = Path(data_dir).expanduser().resolve(strict=False) / "uploads" / "construction_audit_mvp"
    destination = uploads / f"plan_{source_hash[:24]}{suffix}"
    if not destination.exists() or sha256_file(destination) != source_hash:
        _atomic_verified_copy(source, destination, source_hash)
    return destination


def _site_photo_projection(api: Any, source: Path, suffix: str, source_hash: str) -> Path:
    info: dict[str, Any] = {}
    getter = getattr(api, "get_runtime_info", None)
    if callable(getter):
        try:
            candidate = getter()
            info = candidate if isinstance(candidate, dict) else {}
        except Exception:
            info = {}
    data_dir = str(info.get("data_dir") or "").strip()
    if not data_dir:
        return source
    uploads = Path(data_dir).expanduser().resolve(strict=False) / "uploads" / "construction_audit_mvp"
    destination = uploads / f"photo_{source_hash[:24]}{suffix}"
    if not destination.exists() or sha256_file(destination) != source_hash:
        _atomic_verified_copy(source, destination, source_hash)
    return destination


def vision_delegation(plan_id: str, image_ref: str) -> dict[str, Any]:
    source_types = ", ".join(sorted(vision.SOURCE_TYPES))
    context = canonical_json_text({"plan_id": plan_id, "image_ref": image_ref})
    labeled = {
        "value": 1.0,
        "confidence": 1.0,
        "source_type": "explicit_plan_label",
        "evidence_text": "подпись на плане",
    }
    example = canonical_json_text(
        {
            "schema_version": 1,
            "plan_id": plan_id,
            "image_quality": {"usable": True, "issues": []},
            "object_name_suggestion": None,
            "rooms": [
                {
                    "source_room_id": "1",
                    "name": "Офис 1",
                    "floor_area_m2": {**labeled, "source_type": "explicit_area_label"},
                    "length_m": labeled,
                    "width_m": labeled,
                    "perimeter_m": {
                        **labeled,
                        "source_type": "derived_from_explicit_dimensions",
                    },
                    "height_m": labeled,
                    "doors": [
                        {
                            "element_id": "Д-1",
                            "width_m": labeled,
                            "height_m": labeled,
                        }
                    ],
                    "windows": [],
                    "warnings": [],
                }
            ],
            "general_warnings": [],
        }
    )
    return {
        "role": "construction-plan-vision",
        "model_lane": "main",
        "memory_mode": "empty",
        "write_surface": "read_only",
        "objective": (
            f"Открой через view_image файл строительного плана строго по абсолютному пути {image_ref} "
            "из context.image_ref. Используй ровно этот путь и не заменяй его исходным attachment "
            "path. Извлеки помещения, подписанные размеры, площади, высоты, двери и окна. Для каждого "
            "измерения верни value, confidence, source_type и evidence_text."
        ),
        "expected_output": (
            "Верни ровно одну строку без Markdown и пояснений: "
            f"FINAL ANSWER: [BEGIN_SUBTASK_OUTPUT]{example}[END_SUBTASK_OUTPUT]"
        ),
        "context": context,
        "constraints": (
            "Не использовать смету, не подставлять стандартные или предполагаемые значения, "
            "не оценивать размеры по внешнему виду, не изменять файлы и не вызывать tools "
            f"construction_audit_mvp. Единственный допустимый image_ref: {image_ref}. Первым файловым "
            "действием вызови view_image(path=context.image_ref); используй context.image_ref дословно "
            "и не преобразовывай путь. Не заменяй его исходным attachment path. Не используй "
            "artifact_store, attachments, task_results или file://. Не ищи файл через list_files, "
            "search_code, browse_page или vlm_query. Если view_image не сработал, верни ошибку "
            "Vision/orchestration и остановись; не создавай geometry без успешного view_image. Если "
            "значение не найдено: value=null, confidence=0, source_type=not_found, evidence_text=\"\". "
            "image_quality — объект {usable:boolean, issues:string[]}. Каждое помещение содержит "
            "source_room_id, name, floor_area_m2, length_m, width_m, perimeter_m, height_m, doors, "
            "windows и warnings. Каждая дверь и окно используют element_id, width_m и height_m; "
            "размеры — объекты {value, confidence, source_type, evidence_text}. "
            "Дверь является проёмом на границе помещений, а не собственностью одной комнаты. Если "
            "одна дверь расположена на общей границе двух распознанных помещений, обязательно включи "
            "её в doors обоих помещений с одним и тем же element_id и одинаковыми размерами. Для "
            "двери между помещением и внешним либо нераспознанным пространством включи её только в "
            "распознанное помещение. Не выбирай единственную комнату по направлению открывания или "
            "положению подписи. Не определяй по плану, к какой комнате дверь отнесена в смете. "
            f"Допустимые source_type: {source_types}. Не использовать id вместо element_id и не "
            "возвращать image_quality строкой. Завершить ответ точной однострочной формой из "
            "expected_output, начиная с FINAL ANSWER:."
        ),
    }


def mapping_delegation(
    estimate: dict[str, Any],
    geometry: dict[str, Any],
    delegation_token: str,
    price_catalog: dict[str, Any],
) -> dict[str, Any]:
    context = {
        "delegation_token": delegation_token,
        "canonical_rooms": [
            {"room_id": room["room_id"], "name": room["name"]}
            for room in geometry["rooms"]
        ],
        "estimate_rooms": list(estimate["rooms"]),
        "estimate_works": [
            {
                "source_row": row["source_row"],
                "name": row["work_name"],
                "unit": row["unit"],
            }
            for row in estimate["rows"]
            if row.get("work_name")
        ],
        "supported_works": [
            {"name": name, "metric": definition[0], "units": list(definition[1])}
            for name, definition in sorted(SUPPORTED_WORKS.items())
        ],
        "mcp_price_catalog": [
            {
                "id": item["id"],
                "name": item["name"],
                "unit": item["unit"],
                "price": item["price"],
            }
            for item in price_catalog["items"]
        ],
    }
    example = canonical_json_text(
        {
            "schema_version": 3,
            "delegation_token": delegation_token,
            "room_matches": [
                {
                    "estimate_room": "<estimate_rooms item>",
                    "model_room_id": "room_001",
                    "confidence": 1.0,
                    "reason": "<non-empty reason>",
                }
            ],
            "room_unresolved": [],
            "work_matches": [
                {
                    "source_row": 2,
                    "estimate_work": "<estimate_works.name>",
                    "canonical_work": "<supported_works.name>",
                    "confidence": 1.0,
                    "reason": "<non-empty reason>",
                }
            ],
            "work_unsupported": [],
            "work_unresolved": [],
            "price_matches": [
                {
                    "source_row": 2,
                    "estimate_work": "<estimate_works.name>",
                    "mcp_work_id": "<mcp_price_catalog.id>",
                    "confidence": 1.0,
                    "reason": "<non-empty reason>",
                }
            ],
            "price_unsupported": [],
            "price_unresolved": [],
        }
    )
    return {
        "role": "construction-estimate-mapper",
        "model_lane": "light",
        "memory_mode": "empty",
        "write_surface": "read_only",
        "objective": (
            "Независимо сопоставь estimate rooms с canonical rooms, estimate works с supported "
            "works для проверки объёмов и те же строки estimate works с mcp_price_catalog для "
            "проверки стоимости. Учитывай смысл названия, действие, объект работы и единицу "
            "измерения. Для ценового сопоставления верни только стабильный mcp_work_id. Верни "
            "результат строго по контракту expected_output."
        ),
        "expected_output": (
            "Используй ровно указанные ключи. Формы непустых дополнительных элементов: "
            "room_unresolved={estimate_room:string,candidate_room_ids:string[],reason:string,"
            "requires_human_confirmation:boolean}; work_unsupported={source_row:integer,"
            "estimate_work:string,reason:string}; work_unresolved={source_row:integer,"
            "estimate_work:string,candidate_canonical_works:string[],reason:string,"
            "requires_human_confirmation:boolean}; price_unsupported={source_row:integer,"
            "estimate_work:string,reason:string}; price_unresolved={source_row:integer,"
            "estimate_work:string,candidate_mcp_work_ids:string[],reason:string,"
            "requires_human_confirmation:boolean}. Верни ровно одну строку без пояснений и Markdown: "
            f"FINAL ANSWER: [BEGIN_SUBTASK_OUTPUT]{example}[END_SUBTASK_OUTPUT]"
        ),
        "context": canonical_json_text(context),
        "constraints": (
            "Не читать исходные файлы, не рассчитывать количества, не изменять файлы и не "
            "вызывать tools construction_audit_mvp. Не сопоставлять демонтаж с установкой, "
            "поставку с монтажом, ремонт или окраску двери с установкой двери, закупку "
            "материала с устройством конструкции. Вернуть delegation_token из context "
            "дословно и без изменений. model_room_id брать только из canonical_rooms[].room_id. "
            "source_row и estimate_work брать из одного и того же элемента estimate_works; каждый "
            "элемент estimate_works покрыть ровно один раз в work_* и ещё ровно один раз в "
            "price_*. canonical_work — только из supported_works[].name; mcp_work_id — только из "
            "mcp_price_catalog[].id. Не полагаться на точное совпадение названий. Не возвращать, "
            "не изменять и не вычислять price: эталонную цену после маппинга повторно получает "
            "Python по mcp_work_id из сохранённого каталога. Не добавлять другие ключи. "
            "Запрещённые альтернативные поля: canonical_room_id, room_id, canonical_name, metric, "
            "supported_metric, supported_name, estimate_units, matched_unit, matched_units. "
            "Завершить ответ точной однострочной формой из expected_output, начиная с FINAL ANSWER:."
        ),
    }


def deterministic_mapping(
    estimate: dict[str, Any],
    geometry: dict[str, Any],
    delegation_token: str,
    price_catalog: dict[str, Any],
) -> dict[str, Any] | None:
    """Resolve only exact, unit-compatible MVP matches; otherwise defer to LLM."""
    rooms_by_name: dict[str, list[dict[str, Any]]] = {}
    for room in geometry.get("rooms", []):
        rooms_by_name.setdefault(_technical_name(room.get("name", "")), []).append(room)
    room_matches = []
    for estimate_room in estimate.get("rooms", []):
        candidates = rooms_by_name.get(_technical_name(estimate_room), [])
        if len(candidates) != 1:
            return None
        room_matches.append(
            {
                "estimate_room": estimate_room,
                "model_room_id": candidates[0]["room_id"],
                "confidence": 1.0,
                "reason": "Точное совпадение названия помещения.",
            }
        )

    works_by_name = {_technical_name(name): name for name in SUPPORTED_WORKS}
    catalog_by_name: dict[str, list[dict[str, Any]]] = {}
    for item in price_catalog.get("items", []):
        catalog_by_name.setdefault(_technical_name(item.get("name", "")), []).append(item)
    work_matches = []
    price_matches = []
    for row in estimate.get("rows", []):
        work_name = row.get("work_name")
        if not work_name:
            continue
        canonical_work = works_by_name.get(_technical_name(work_name))
        if canonical_work is None:
            return None
        allowed_units = {_technical_name(unit) for unit in SUPPORTED_WORKS[canonical_work][1]}
        if _technical_name(row.get("unit", "")) not in allowed_units:
            return None
        catalog_candidates = [
            item
            for item in catalog_by_name.get(_technical_name(work_name), [])
            if _technical_name(item.get("unit", "")) == _technical_name(row.get("unit", ""))
        ]
        if len(catalog_candidates) != 1:
            return None
        work_matches.append(
            {
                "source_row": row["source_row"],
                "estimate_work": work_name,
                "canonical_work": canonical_work,
                "confidence": 1.0,
                "reason": "Точное совпадение работы и совместимая единица.",
            }
        )
        price_matches.append(
            {
                "source_row": row["source_row"],
                "estimate_work": work_name,
                "mcp_work_id": catalog_candidates[0]["id"],
                "confidence": 1.0,
                "reason": "Точное совпадение работы и единицы каталога.",
            }
        )
    return {
        "schema_version": 3,
        "delegation_token": delegation_token,
        "room_matches": room_matches,
        "room_unresolved": [],
        "work_matches": work_matches,
        "work_unsupported": [],
        "work_unresolved": [],
        "price_matches": price_matches,
        "price_unsupported": [],
        "price_unresolved": [],
    }


def _decode_mcp_catalog_projection(response: Any) -> tuple[Any, str]:
    if (
        isinstance(response, dict)
        and isinstance(response.get("runtime_text_projection"), str)
        and set(response) <= {"result", "runtime_text_projection"}
    ):
        return _decode_mcp_catalog_projection(response["runtime_text_projection"])
    if not isinstance(response, str):
        return response, "structured_object"
    if len(response) > 2_000_000:
        raise AuditError("price_catalog_schema_invalid", "Текстовый ответ MCP слишком велик.")
    # Runtime иногда проецирует массив result как несколько JSON-объектов подряд.
    # Разбираем только такие законченные значения; содержимое полей не «чинится».
    decoder = json.JSONDecoder()
    values: list[Any] = []
    position = 0
    while position < len(response):
        starts = [index for index in (response.find("{", position), response.find("[", position)) if index >= 0]
        if not starts:
            break
        start = min(starts)
        try:
            value, end = decoder.raw_decode(response, start)
        except json.JSONDecodeError:
            position = start + 1
            continue
        values.append(value)
        position = end
    if len(values) == 1 and isinstance(values[0], dict) and "result" in values[0]:
        return values[0], "runtime_text_projection"
    if values and all(isinstance(item, dict) for item in values):
        return {"result": values}, "runtime_text_projection"
    return response, "runtime_text_projection"


def validate_price_catalog_response(response: Any) -> dict[str, Any]:
    """Validate an MCP wrapper or its deterministic runtime text projection."""
    response, transport = _decode_mcp_catalog_projection(response)
    if not isinstance(response, dict):
        raise AuditError(
            "price_catalog_schema_invalid",
            "Ответ MCP должен быть object с массивом result.",
            {"validation_errors": [{"path": "", "reason": "must be an object"}]},
        )
    if "result" not in response:
        raise AuditError(
            "price_catalog_schema_invalid",
            "В ответе MCP отсутствует обязательное поле result.",
            {"validation_errors": [{"path": "result", "reason": "required field is missing"}]},
        )
    raw_items = response.get("result")
    if not isinstance(raw_items, list):
        raise AuditError(
            "price_catalog_schema_invalid",
            "Поле result в ответе MCP должно быть массивом.",
            {"validation_errors": [{"path": "result", "reason": "must be an array"}]},
        )
    if len(raw_items) > PRICE_CATALOG_MAX_ITEMS:
        raise AuditError(
            "price_catalog_schema_invalid",
            f"Каталог MCP не должен содержать более {PRICE_CATALOG_MAX_ITEMS} позиций.",
            {"validation_errors": [{"path": "result", "reason": "too many items"}]},
        )

    errors: list[dict[str, str]] = []
    normalized: list[dict[str, str]] = []
    seen_ids: set[str] = set()
    for index, raw in enumerate(raw_items):
        path = f"result[{index}]"
        if not isinstance(raw, dict):
            errors.append({"path": path, "reason": "must be an object"})
            continue
        missing = sorted({"id", "name", "unit", "price"} - set(raw))
        for field in missing:
            errors.append({"path": f"{path}.{field}", "reason": "required field is missing"})

        clean_text: dict[str, str] = {}
        for field in ("id", "name", "unit"):
            value = raw.get(field)
            if not isinstance(value, str) or not value.strip() or len(value.strip()) > 200:
                errors.append({"path": f"{path}.{field}", "reason": "must be a non-empty string up to 200 characters"})
            else:
                clean_text[field] = value.strip()

        price_value = raw.get("price")
        price: Decimal | None = None
        if isinstance(price_value, bool) or not isinstance(price_value, (int, float, Decimal)):
            errors.append({"path": f"{path}.price", "reason": "must be a finite non-negative number"})
        else:
            price = _decimal(price_value)
            if price is None or price < 0:
                errors.append({"path": f"{path}.price", "reason": "must be a finite non-negative number"})

        item_id = clean_text.get("id")
        if item_id:
            if item_id in seen_ids:
                errors.append({"path": f"{path}.id", "reason": "duplicate id"})
            else:
                seen_ids.add(item_id)
        if not missing and len(clean_text) == 3 and price is not None and price >= 0:
            normalized.append({**clean_text, "price": decimal_text(price) or "0"})

    if errors:
        raise AuditError(
            "price_catalog_schema_invalid",
            "Каталог работ MCP не прошёл validation.",
            {"validation_errors": errors},
        )
    return {
        "schema_version": 1,
        "source_tool": PRICE_CATALOG_TOOL,
        "source_transport": transport,
        "items": normalized,
    }


def save_price_catalog(
    api: Any,
    job_id: str,
    catalog_response: Any,
) -> tuple[dict[str, Any], dict[str, Any]]:
    root = job_dir(api, job_id)
    manifest = load_manifest(root)
    if not manifest.get("geometry_confirmed"):
        raise AuditError(
            "geometry_confirmation_required",
            "Сначала подтвердите geometry отдельным новым сообщением пользователя.",
        )
    geometry = read_json(output_path(root, "geometry.json"), code="geometry_required")
    if geometry_sha256(geometry) != manifest.get("confirmed_geometry_sha256"):
        raise AuditError("geometry_changed_since_confirmation", "Подтверждённый hash geometry не совпадает с текущим.")
    if manifest.get("geometry_confirmed_revision") != manifest.get("geometry_revision"):
        raise AuditError("stale_geometry_confirmation", "Подтверждение относится не к текущей geometry revision.")

    catalog = validate_price_catalog_response(catalog_response)
    _invalidate_audit(manifest, root)
    fingerprint = canonical_json_sha256(catalog)
    atomic_write_json(output_path(root, "price_catalog.json"), catalog)
    manifest["price_catalog_sha256"] = fingerprint
    manifest["price_catalog_items_count"] = len(catalog["items"])
    manifest["price_catalog_source_tool"] = PRICE_CATALOG_TOOL
    manifest["price_catalog_source_transport"] = catalog["source_transport"]
    manifest["price_catalog_validation_status"] = "validated"
    manifest["price_catalog_saved_at"] = now_iso()
    manifest["skill_version"] = SKILL_VERSION
    if manifest.get("input_mode") == "plan_only":
        manifest.pop("mapping_generation_token", None)
        save_manifest(root, manifest)
        return manifest, {
            "status": "estimate_generation_offer",
            "catalog_items_count": len(catalog["items"]),
            "price_catalog_sha256": fingerprint,
            "source_transport": catalog["source_transport"],
            "offer_markdown": (
                "Геометрия подтверждена, каталог цен MCP сохранён. Могу сформировать "
                "предварительную XLSX-смету по всем поддерживаемым работам, для которых "
                "достаточно геометрических данных. Сформировать смету?"
            ),
            "next_action": "await_estimate_generation_confirmation",
            "assistant_instruction": ESTIMATE_OFFER_ASSISTANT_INSTRUCTION,
        }

    # Токен появляется только после validation каталога: Mapping должен быть
    # привязан сразу к geometry revision и к фактическому набору внешних работ.
    manifest["mapping_generation_token"] = secrets.token_urlsafe(32)
    save_manifest(root, manifest)
    estimate = read_json(output_path(root, "estimate_normalized.json"), code="estimate_required")
    fast_mapping = deterministic_mapping(
        estimate,
        geometry,
        manifest["mapping_generation_token"],
        catalog,
    )
    # Модель подключаем только там, где требуется понять формулировку. Точные
    # совпадения имён и единиц надёжнее и дешевле закрыть детерминированно.
    if fast_mapping is not None:
        return manifest, {
            "status": "price_catalog_saved",
            "catalog_items_count": len(catalog["items"]),
            "price_catalog_sha256": fingerprint,
            "source_transport": catalog["source_transport"],
            "mapping_mode": "deterministic_exact_match",
            "mapping_task_id": "deterministic_mapping",
            "mapping": fast_mapping,
            "next_action": "run_audit",
        }
    delegation = mapping_delegation(
        estimate,
        geometry,
        manifest["mapping_generation_token"],
        catalog,
    )
    return manifest, {
        "status": "price_catalog_saved",
        "catalog_items_count": len(catalog["items"]),
        "price_catalog_sha256": fingerprint,
        "source_transport": catalog["source_transport"],
        "mapping_mode": "subagent_required",
        "mapping_delegation": delegation,
        "next_action": "schedule_mapping_subagent",
    }


def import_documents(
    api: Any,
    job_id: str,
    estimate_descriptor: Any,
    plan_descriptor: Any,
) -> tuple[dict[str, Any], dict[str, Any]]:
    root = job_dir(api, job_id)
    manifest = load_manifest(root)
    estimate_source, estimate_name, _ = _validate_attachment_descriptor(estimate_descriptor, {".xlsx"})
    plan_source, plan_name, plan_suffix = _validate_attachment_descriptor(plan_descriptor, {".png", ".jpg", ".jpeg"})
    _validate_xlsx(estimate_source)
    plan_mime = _image_mime(plan_source, plan_suffix)
    normalized = parse_estimate(estimate_source)
    estimate_hash = sha256_file(estimate_source)
    plan_hash = sha256_file(plan_source)
    old_documents = manifest.get("documents") if isinstance(manifest.get("documents"), dict) else {}
    estimate_changed = (old_documents.get("estimate") or {}).get("sha256") != estimate_hash
    plan_changed = (old_documents.get("plan") or {}).get("sha256") != plan_hash

    _atomic_verified_copy(estimate_source, root / "assets" / "estimate.xlsx", estimate_hash)
    _atomic_verified_copy(plan_source, root / "assets" / f"plan{plan_suffix}", plan_hash)
    projection = _vision_projection(api, plan_source, plan_suffix, plan_hash)
    if projection.parent.name != "construction_audit_mvp" or projection.parent.parent.name != "uploads":
        raise AuditError(
            "vision_projection_unavailable",
            "Не удалось создать разрешённый upload projection для Vision-subagent.",
        )
    projection_path = str(projection)
    # Смета не влияет на распознанный план, поэтому при её замене сбрасываем
    # только аудит. Новый план делает неактуальной уже саму geometry.
    if estimate_changed:
        _invalidate_audit(manifest, root)
    if plan_changed:
        _invalidate_geometry(manifest, root)
    atomic_write_json(output_path(root, "estimate_normalized.json"), normalized)
    manifest["documents_imported"] = True
    manifest["input_mode"] = "estimate_and_plan"
    manifest["documents"] = {
        "estimate": {
            "filename": estimate_name,
            "sha256": estimate_hash,
            "size_bytes": estimate_source.stat().st_size,
            "imported_name": "estimate.xlsx",
        },
        "plan": {
            "filename": plan_name,
            "sha256": plan_hash,
            "size_bytes": plan_source.stat().st_size,
            "mime": plan_mime,
            "plan_id": vision.PLAN_ID,
            "imported_name": f"plan{plan_suffix}",
            "vision_source_path": projection_path,
        },
    }
    save_manifest(root, manifest)
    return manifest, {
        "status": "documents_imported",
        "estimate": {
            "rows_count": len(normalized["rows"]),
            "rooms": normalized["rooms"],
            "works": normalized["works"],
        },
        "plan": {"plan_id": vision.PLAN_ID, "source_path": projection_path},
        "next_action": "schedule_vision_subagent",
        "vision_delegation": vision_delegation(vision.PLAN_ID, projection_path),
    }


def import_plan(
    api: Any,
    job_id: str,
    plan_descriptor: Any,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Import only a plan without changing the existing XLSX audit import path."""
    root = job_dir(api, job_id)
    manifest = load_manifest(root)
    plan_source, plan_name, plan_suffix = _validate_attachment_descriptor(
        plan_descriptor, {".png", ".jpg", ".jpeg"}
    )
    plan_mime = _image_mime(plan_source, plan_suffix)
    plan_hash = sha256_file(plan_source)
    old_documents = manifest.get("documents") if isinstance(manifest.get("documents"), dict) else {}
    plan_changed = (old_documents.get("plan") or {}).get("sha256") != plan_hash

    _atomic_verified_copy(plan_source, root / "assets" / f"plan{plan_suffix}", plan_hash)
    projection = _vision_projection(api, plan_source, plan_suffix, plan_hash)
    if projection.parent.name != "construction_audit_mvp" or projection.parent.parent.name != "uploads":
        raise AuditError(
            "vision_projection_unavailable",
            "Не удалось создать разрешённый upload projection для Vision-subagent.",
        )
    if plan_changed:
        _invalidate_geometry(manifest, root)
    else:
        _invalidate_audit(manifest, root)
    output_path(root, "estimate_normalized.json").unlink(missing_ok=True)
    (root / "assets" / "estimate.xlsx").unlink(missing_ok=True)

    projection_path = str(projection)
    manifest["documents_imported"] = True
    manifest["input_mode"] = "plan_only"
    manifest["documents"] = {
        "plan": {
            "filename": plan_name,
            "sha256": plan_hash,
            "size_bytes": plan_source.stat().st_size,
            "mime": plan_mime,
            "plan_id": vision.PLAN_ID,
            "imported_name": f"plan{plan_suffix}",
            "vision_source_path": projection_path,
        },
    }
    save_manifest(root, manifest)
    return manifest, {
        "status": "plan_imported",
        "plan": {"plan_id": vision.PLAN_ID, "source_path": projection_path},
        "next_action": "schedule_vision_subagent",
        "vision_delegation": vision_delegation(vision.PLAN_ID, projection_path),
    }


def _ctx_value(ctx: Any, name: str) -> str:
    value = getattr(ctx, name, "") if ctx is not None else ""
    return str(value or "").strip()


def _client_message_id(ctx: Any) -> str:
    metadata = getattr(ctx, "task_metadata", {}) if ctx is not None else {}
    if not isinstance(metadata, dict):
        return ""
    return str(metadata.get("client_message_id") or "").strip()


def _markdown_cell(value: Any) -> str:
    return str(value if value not in (None, "") else "—").replace("|", "\\|").replace("\n", " ")


def _opening_summary(openings: list[dict[str, Any]]) -> str:
    if not openings:
        return "нет"
    return "; ".join(
        f"{item['element_id']} — {item['width_m']} × {item['height_m']} м"
        for item in openings
    )


GEOMETRY_FIELD_LABELS = {
    "floor_area_m2": "площадь пола",
    "length_m": "длина",
    "width_m": "ширина",
    "perimeter_m": "периметр",
    "height_m": "высота",
}


def _review_room_labels(rooms: list[dict[str, Any]]) -> dict[str, str]:
    """Use room names in the user view and add technical IDs only for duplicates."""
    name_counts: dict[str, int] = {}
    for room in rooms:
        name = str(room.get("name") or "Помещение").strip()
        key = name.casefold()
        name_counts[key] = name_counts.get(key, 0) + 1
    return {
        str(room.get("room_id") or ""): (
            f"{str(room.get('name') or 'Помещение').strip()} (`{room.get('room_id')}`)"
            if name_counts[str(room.get("name") or "Помещение").strip().casefold()] > 1
            else str(room.get("name") or "Помещение").strip()
        )
        for room in rooms
    }


def _review_field_label(field: Any) -> str:
    value = str(field or "параметр")
    parts = value.split(".")
    if len(parts) == 3 and parts[0] in {"doors", "windows"}:
        opening = "дверной проём" if parts[0] == "doors" else "окно"
        return f"{opening} {parts[1]} — {GEOMETRY_FIELD_LABELS.get(parts[2], parts[2])}"
    return GEOMETRY_FIELD_LABELS.get(value, value)


def _room_dimensions_cell(room: dict[str, Any]) -> str:
    length = str(room.get("length_m") or "—")
    width = str(room.get("width_m") or "—")
    if length == "—" and width == "—":
        return "**нужно уточнить**"
    if length == "—":
        length = "**уточнить**"
    if width == "—":
        width = "**уточнить**"
    return f"{length} × {width}"


def _derived_review_value(room: dict[str, Any], field: str) -> str:
    value = str(room.get(field) or "—")
    if value == "—" and (str(room.get("length_m") or "—") == "—" or str(room.get("width_m") or "—") == "—"):
        return "рассчитается"
    return value


def _missing_room_guidance(room: dict[str, Any], room_label: str) -> list[str]:
    missing = {str(item) for item in room.get("missing_fields", [])}
    messages: list[str] = []
    dimensions = [GEOMETRY_FIELD_LABELS[field] for field in ("length_m", "width_m") if field in missing]
    if dimensions:
        if len(dimensions) == 2:
            requested = "длину и ширину"
        else:
            requested = dimensions[0]
        derived = [
            GEOMETRY_FIELD_LABELS[field]
            for field in ("floor_area_m2", "perimeter_m")
            if field in missing
        ]
        suffix = ""
        if derived:
            verb = "рассчитается" if len(derived) == 1 else "рассчитаются"
            suffix = f"; {' и '.join(derived)} {verb} автоматически"
        messages.append(f"**{room_label}:** укажите {requested}{suffix}.")
        missing.difference_update({"length_m", "width_m", "floor_area_m2", "perimeter_m"})
    for field in sorted(missing):
        messages.append(f"**{room_label}:** укажите {_review_field_label(field)}.")
    return messages


def _require_consistent_shared_doors(geometry: dict[str, Any]) -> None:
    seen: dict[str, tuple[Any, Any, str]] = {}
    conflicts: list[dict[str, Any]] = []
    for room in geometry.get("rooms", []):
        room_id = str(room.get("room_id") or "")
        for door in room.get("doors", []):
            element_id = str(door.get("element_id") or "")
            current = (door.get("width_m"), door.get("height_m"), room_id)
            previous = seen.get(element_id)
            if previous is not None and previous[:2] != current[:2]:
                conflicts.append(
                    {
                        "element_id": element_id,
                        "room_ids": [previous[2], current[2]],
                        "dimensions": [
                            {"width_m": previous[0], "height_m": previous[1]},
                            {"width_m": current[0], "height_m": current[1]},
                        ],
                    }
                )
            else:
                seen[element_id] = current
    if conflicts:
        raise AuditError(
            "shared_door_dimension_conflict",
            "Одна и та же дверь имеет разные размеры в смежных помещениях; geometry нельзя подтвердить или использовать в аудите.",
            {"doors": conflicts},
        )


def _review_issue_text(item: Any, room_labels: dict[str, str] | None = None) -> str:
    if not isinstance(item, dict):
        return str(item)
    labels = room_labels or {}
    if item.get("type") == "geometry_arithmetic_conflict":
        room = labels.get(str(item.get("room_id")), str(item.get("room_id") or "помещение"))
        return (
            f"{room}: {_review_field_label(item.get('field'))} — "
            f"указано {item.get('reported_value', '—')}, расчётное значение "
            f"{item.get('calculated_value', '—')}"
        )
    if item.get("type") == "opening_dimension_conflict":
        rooms = ", ".join(labels.get(str(room_id), str(room_id)) for room_id in (item.get("room_ids") or []))
        return f"{item.get('element_id', 'проём')}: разные размеры в помещениях {rooms}"
    if "room_id" in item and "field" in item:
        room = labels.get(str(item["room_id"]), str(item["room_id"]))
        return f"{room}: отсутствует {_review_field_label(item['field'])}"
    return canonical_json_text(item)


def build_geometry_review_markdown(
    review: dict[str, Any],
    geometry_revision: int,
) -> str:
    """Format only fields already present in the structured geometry review."""
    image_quality = review["image_quality"]
    quality = "план пригоден для анализа" if image_quality["usable"] else "план непригоден для анализа"
    room_labels = _review_room_labels(review["rooms"])
    lines = [
        "# Проверка геометрии",
        "",
        f"**Объект:** {review['object_name']}  ",
        f"**Версия:** {geometry_revision}  ",
        f"**Качество:** {quality}",
        "",
        "## Помещения",
        "",
        "| Помещение | Размеры, м (Д × Ш) | Площадь, м² | Периметр, м | Высота, м |",
        "|---|---:|---:|---:|---:|",
    ]
    for room in review["rooms"]:
        lines.append(
            "| " + " | ".join(
                (
                    _markdown_cell(room_labels[str(room["room_id"])]),
                    _room_dimensions_cell(room),
                    _derived_review_value(room, "floor_area_m2"),
                    _derived_review_value(room, "perimeter_m"),
                    _markdown_cell(room["height_m"]),
                )
            ) + " |"
        )

    lines.extend(
        [
            "",
            "## Двери и окна",
            "",
            "Один межкомнатный дверной проём показывается у каждого смежного помещения с одинаковым ID; в итоге объекта он считается один раз.",
        ]
    )
    for room in review["rooms"]:
        lines.extend(
            [
                "",
                f"### {room_labels[str(room['room_id'])]}",
                f"- **Дверные проёмы:** {_opening_summary(room['doors'])}",
                f"- **Окна:** {_opening_summary(room['windows'])}",
            ]
        )

    totals = review["totals"]
    lines.extend(
        [
            "",
            "## Итоги",
            "",
            "| Помещений | Общая площадь, м² | Уникальных дверей | Уникальных окон |",
            "|---:|---:|---:|---:|",
            (
                f"| {totals['rooms_count']} | {totals['floor_area_m2']} | "
                f"{totals['unique_doors_count']} | {totals['unique_windows_count']} |"
            ),
        ]
    )

    quality_issues = list(image_quality["issues"])
    warnings = list(dict.fromkeys(str(item) for item in review["warnings"]))
    issues = [*quality_issues, *warnings]
    lines.extend(["", "## Требуют внимания", ""])
    lines.extend(f"- {item}" for item in issues)
    if not issues:
        lines.append("- нет")

    lines.extend(["", "## Что нужно уточнить", ""])
    missing_guidance = [
        message
        for room in review["rooms"]
        for message in _missing_room_guidance(room, room_labels[str(room["room_id"])])
    ]
    lines.extend(f"- {message}" for message in missing_guidance)
    if not missing_guidance:
        lines.append("- Данных достаточно для расчёта.")

    lines.extend(["", "## Конфликты", ""])
    lines.extend(f"- {_review_issue_text(item, room_labels)}" for item in review["conflicts"])
    if not review["conflicts"]:
        lines.append("- нет")

    lines.extend(
        [
            "",
            "## Подтверждение",
            "",
            (
                "Для продолжения пользователь должен отправить отдельное новое входящее сообщение "
                "с подтверждением — например, **«Геометрия верна»** — или перечислить исправления."
            ),
            "",
            GEOMETRY_REVIEW_TERMINAL_MARKER,
        ]
    )
    return "\n".join(lines)


def render_geometry_review(
    api: Any,
    job_id: str,
    geometry_revision: Any,
) -> dict[str, Any]:
    root = Path(api.skill_job_dir(validate_job_id(job_id))).resolve(strict=False)
    manifest = load_manifest(root)
    # Renderer может быть вызван после handoff с задержкой. За это время правка
    # пользователя уже могла создать следующую revision.
    if type(geometry_revision) is not int or geometry_revision != manifest.get("geometry_revision"):
        raise AuditError("stale_geometry_revision", "Запрошенная версия геометрии устарела.")
    review = read_json(output_path(root, "geometry_review.json"), code="geometry_required")
    if not isinstance(review, dict):
        raise AuditError("storage_error", "Сохранённый review геометрии повреждён.")
    return {
        "status": "geometry_review_required",
        "geometry_revision": geometry_revision,
        "review_markdown": build_geometry_review_markdown(review, geometry_revision),
        "next_action": "display_review_markdown_verbatim",
        "assistant_instruction": GEOMETRY_REVIEW_ASSISTANT_INSTRUCTION,
    }


def save_geometry(
    api: Any,
    ctx: Any,
    job_id: str,
    vision_task_id: Any,
    analysis: Any,
) -> tuple[dict[str, Any], dict[str, Any]]:
    root = job_dir(api, job_id)
    manifest = load_manifest(root)
    plan = (manifest.get("documents") or {}).get("plan")
    if not manifest.get("documents_imported") or not isinstance(plan, dict):
        raise AuditError("plan_required", "Сначала импортируйте план.")
    if not isinstance(vision_task_id, str) or not vision_task_id.strip() or len(vision_task_id.strip()) > 200:
        raise AuditError("geometry_analysis_invalid", "vision_task_id должен быть непустой строкой до 200 символов.")
    normalized_task_id = vision_task_id.strip()
    parent_task_id = _ctx_value(ctx, "task_id")
    if parent_task_id and normalized_task_id == parent_task_id:
        raise AuditError(
            "vision_subagent_required",
            "Geometry разрешено сохранять только из результата отдельного Vision-субагента.",
            {"next_action": "schedule_vision_subagent"},
        )
    attempts = int(manifest.get("vision_attempts") or 0)
    if attempts >= 2:
        raise AuditError("geometry_analysis_failed", "Не удалось получить валидный анализ плана после двух попыток.")
    # Сначала проверяем весь модельный ответ. При schema error последний валидный
    # geometry artifact остаётся на месте и может быть использован для диагностики.
    try:
        validated = vision.validate_analysis(analysis, str(plan.get("plan_id") or vision.PLAN_ID))
    except vision.GeometryValidationError as exc:
        manifest["vision_attempts"] = attempts + 1
        code = "geometry_analysis_failed" if manifest["vision_attempts"] >= 2 else "geometry_schema_invalid"
        message = (
            "Не удалось получить валидный анализ плана после двух попыток."
            if code == "geometry_analysis_failed"
            else "Результат Vision-субагента не прошёл validation; разрешена одна повторная попытка."
        )
        save_manifest(root, manifest)
        raise AuditError(
            code,
            message,
            {
                "validation_errors": exc.errors,
                "allowed_next_action": "retry_vision" if manifest["vision_attempts"] < 2 else "stop",
                "vision_attempt": manifest["vision_attempts"],
            },
        ) from exc

    manifest["vision_attempts"] = attempts + 1
    geometry = vision.canonicalize_geometry(validated, manifest["object_name"])
    geometry_fingerprint = geometry_sha256(geometry)
    review = vision.build_geometry_review(geometry)
    revision = int(manifest.get("geometry_revision") or 0) + 1
    confirmation_question = GEOMETRY_CONFIRMATION_NOTICE
    _invalidate_audit(manifest, root)
    atomic_write_json(output_path(root, "geometry.json"), geometry)
    atomic_write_json(output_path(root, "geometry_review.json"), review)
    manifest["geometry_revision"] = revision
    manifest["geometry_confirmed"] = False
    manifest["vision_task_id"] = normalized_task_id
    manifest["geometry_sha256"] = geometry_fingerprint
    manifest["geometry_schema_version"] = validated["schema_version"]
    manifest["geometry_validation_status"] = "validated"
    manifest["review_task_id"] = _ctx_value(ctx, "task_id")
    client_message_id = _client_message_id(ctx)
    if client_message_id:
        manifest["review_client_message_id"] = client_message_id
    else:
        manifest.pop("review_client_message_id", None)
    for key in (
        "confirmation_task_id",
        "confirmation_client_message_id",
        "confirmed_geometry_sha256",
        "geometry_confirmed_revision",
        "geometry_confirmed_at",
        "mapping_generation_token",
    ):
        manifest.pop(key, None)
    save_manifest(root, manifest)
    return manifest, {
        "status": "geometry_review_required",
        "geometry_revision": revision,
        "confirmation_question": confirmation_question,
        "handoff_barrier_draft": GEOMETRY_HANDOFF_BARRIER_DRAFT,
        "next_action": "cross_subagent_handoff_barrier_then_render_geometry_review",
        "assistant_instruction": GEOMETRY_HANDOFF_BARRIER_ASSISTANT_INSTRUCTION,
    }


def geometry_sha256(geometry: dict[str, Any]) -> str:
    return canonical_json_sha256(geometry)


def _require_new_geometry_turn(
    manifest: dict[str, Any],
    ctx: Any,
    *,
    code: str,
    message: str,
) -> None:
    # Review нельзя подтвердить тем же ответом, которым он был сформирован:
    # нужен отдельный пользовательский ввод, а не решение основной LLM за него.
    current_task_id = _ctx_value(ctx, "task_id")
    review_task_id = str(manifest.get("review_task_id") or "").strip()
    if current_task_id and review_task_id and current_task_id == review_task_id:
        raise AuditError(code, message)
    current_client_id = _client_message_id(ctx)
    review_client_id = str(manifest.get("review_client_message_id") or "").strip()
    if current_client_id and review_client_id and current_client_id == review_client_id:
        raise AuditError(code, message)


def _correction_selector(
    value: Any,
    available: set[str],
    field: str,
) -> set[str]:
    if value == "all":
        return set(available)
    if (
        not isinstance(value, list)
        or not 1 <= len(value) <= 200
        or any(not isinstance(item, str) or not item or len(item) > 200 for item in value)
        or len(value) != len(set(value))
    ):
        raise AuditError(
            "geometry_correction_invalid",
            f"{field} должен быть строкой 'all' или непустым массивом уникальных ID.",
        )
    unknown = sorted(set(value) - available)
    if unknown:
        raise AuditError(
            "geometry_correction_target_not_found",
            f"В {field} указаны неизвестные ID.",
            {"field": field, "unknown_ids": unknown},
        )
    return set(value)


def _room_correction_selector(
    value: Any,
    rooms: list[dict[str, Any]],
    field: str,
) -> set[str]:
    """Resolve stable IDs, source IDs, or an unambiguous exact room name."""
    room_ids = {str(room["room_id"]) for room in rooms}
    if value == "all":
        return room_ids
    if (
        not isinstance(value, list)
        or not 1 <= len(value) <= 200
        or any(not isinstance(item, str) or not item.strip() or len(item) > 200 for item in value)
        or len(value) != len(set(value))
    ):
        raise AuditError(
            "geometry_correction_invalid",
            f"{field} должен быть строкой 'all' или непустым массивом уникальных идентификаторов/названий.",
        )
    aliases: dict[str, set[str]] = {}
    for room in rooms:
        room_id = str(room["room_id"])
        for alias in (room_id, room.get("source_room_id"), room.get("name")):
            if isinstance(alias, str) and alias.strip():
                aliases.setdefault(_technical_name(alias), set()).add(room_id)
    selected: set[str] = set()
    unknown: list[str] = []
    ambiguous: list[str] = []
    for raw in value:
        matches = aliases.get(_technical_name(raw), set())
        if not matches:
            unknown.append(raw)
        elif len(matches) > 1:
            ambiguous.append(raw)
        else:
            selected.update(matches)
    if ambiguous:
        raise AuditError(
            "geometry_correction_target_ambiguous",
            "Название помещения совпадает с несколькими помещениями; укажите room_NNN.",
            {"field": field, "ambiguous_selectors": ambiguous},
        )
    if unknown:
        raise AuditError(
            "geometry_correction_target_not_found",
            f"В {field} указаны неизвестные идентификаторы или названия.",
            {
                "field": field,
                "unknown_ids": unknown,
                "available_rooms": [
                    {
                        "room_id": room.get("room_id"),
                        "source_room_id": room.get("source_room_id"),
                        "name": room.get("name"),
                    }
                    for room in rooms
                ],
            },
        )
    return selected


def _remove_stale_geometry_warnings(
    geometry: dict[str, Any],
    corrected: list[dict[str, Any]],
) -> None:
    field_tokens = {
        "floor_area_m2": ("площад", "floor_area"),
        "length_m": ("длин", "length"),
        "width_m": ("ширин", "width"),
        "perimeter_m": ("периметр", "perimeter"),
        "height_m": ("высот", "height"),
    }
    corrected_by_room: dict[str, set[str]] = {}
    for item in corrected:
        for path in item.get("applied_targets", []):
            parts = str(path).split(".")
            if len(parts) >= 2:
                corrected_by_room.setdefault(parts[0], set()).add(parts[-1])

    # Удаляем предупреждения только по исправленным полям. Соседняя проблема
    # в той же комнате после одной пользовательской правки не исчезает.
    corrected_fields = {
        field
        for fields in corrected_by_room.values()
        for field in fields
    }
    global_tokens = tuple(
        token
        for field in corrected_fields
        for token in field_tokens.get(field, ())
    )

    removed: set[str] = set()
    for room in geometry.get("rooms", []):
        fields = corrected_by_room.get(str(room.get("room_id")), set())
        tokens = tuple(token for field in fields for token in field_tokens.get(field, ()))
        if not tokens:
            continue
        kept = []
        for warning in room.get("warnings", []):
            normalized = str(warning).casefold()
            if any(token in normalized for token in tokens):
                removed.add(str(warning))
            else:
                kept.append(warning)
        room["warnings"] = kept
    geometry["warnings"] = [
        warning
        for warning in geometry.get("warnings", [])
        if str(warning) not in removed
        and not any(token in str(warning).casefold() for token in global_tokens)
    ]


def _correction_value(value: Any, *, maximum: Decimal) -> str | None:
    if value is None:
        return None
    number = _decimal(value)
    if number is None or number <= 0 or number > maximum:
        raise AuditError(
            "geometry_correction_invalid",
            f"Значение правки должно быть больше 0 и не больше {maximum} либо null.",
        )
    return vision.decimal_text(number)


def revise_geometry(
    api: Any,
    ctx: Any,
    job_id: str,
    geometry_revision: Any,
    corrections: Any,
    user_statement: Any,
) -> tuple[dict[str, Any], dict[str, Any]]:
    root = job_dir(api, job_id)
    manifest = load_manifest(root)
    geometry_path = output_path(root, "geometry.json")
    if not geometry_path.exists():
        raise AuditError("geometry_required", "Сначала сохраните geometry и покажите полный review.")
    if type(geometry_revision) is not int or geometry_revision != manifest.get("geometry_revision"):
        raise AuditError("stale_geometry_revision", "Показанная версия геометрии устарела.")
    _require_new_geometry_turn(
        manifest,
        ctx,
        code="correction_requires_new_turn",
        message="Исправление geometry должно поступить отдельным новым сообщением пользователя.",
    )
    if not isinstance(user_statement, str) or not 1 <= len(user_statement.strip()) <= 1000:
        raise AuditError(
            "geometry_correction_invalid",
            "user_statement должен дословно содержать явную правку пользователя длиной до 1000 символов.",
        )
    statement = user_statement
    if not isinstance(corrections, list) or not 1 <= len(corrections) <= GEOMETRY_CORRECTION_MAX_ITEMS:
        raise AuditError(
            "geometry_correction_invalid",
            f"corrections должен содержать от 1 до {GEOMETRY_CORRECTION_MAX_ITEMS} правок.",
        )

    geometry = read_json(geometry_path, code="geometry_required")
    if not isinstance(geometry, dict) or not isinstance(geometry.get("rooms"), list):
        raise AuditError("storage_error", "Сохранённая geometry повреждена.")
    previous_hash = geometry_sha256(geometry)
    rooms = geometry["rooms"]
    room_by_id = {
        room.get("room_id"): room
        for room in rooms
        if isinstance(room, dict) and isinstance(room.get("room_id"), str)
    }
    if len(room_by_id) != len(rooms):
        raise AuditError("storage_error", "Сохранённая geometry содержит некорректные room_id.")

    normalized: list[dict[str, Any]] = []
    changed_paths: set[str] = set()
    evidence = statement[:500]
    for index, raw in enumerate(corrections):
        required = {"target", "room_ids", "element_ids", "field", "value"}
        if not isinstance(raw, dict) or set(raw) != required:
            raise AuditError(
                "geometry_correction_invalid",
                f"corrections[{index}] должен содержать только target, room_ids, element_ids, field и value.",
            )
        target = raw.get("target")
        field = raw.get("field")
        if target not in GEOMETRY_CORRECTION_FIELDS:
            raise AuditError("geometry_correction_invalid", f"corrections[{index}].target не поддерживается.")
        if field not in GEOMETRY_CORRECTION_FIELDS[target]:
            raise AuditError(
                "geometry_correction_invalid",
                f"Поле {field!r} недопустимо для target={target!r}.",
            )
        selected_room_ids = _room_correction_selector(
            raw.get("room_ids"), rooms, f"corrections[{index}].room_ids"
        )
        if not selected_room_ids:
            raise AuditError("geometry_correction_target_not_found", "Geometry не содержит выбранных помещений.")
        maximum = (
            vision.MEASUREMENT_LIMITS[field]
            if target == "rooms"
            else Decimal("100")
        )
        value = _correction_value(raw.get("value"), maximum=maximum)
        measurement = {
            "value": value,
            "confidence": 1.0 if value is not None else 0.0,
            "source_type": "user_correction" if value is not None else "not_found",
            "evidence_text": evidence if value is not None else "",
        }
        applied_targets: list[str] = []

        if target == "rooms":
            if raw.get("element_ids") != "all":
                raise AuditError(
                    "geometry_correction_invalid",
                    f"corrections[{index}].element_ids должен быть 'all' для target='rooms'.",
                )
            for room_id in sorted(selected_room_ids):
                path = f"{room_id}.{field}"
                if path in changed_paths:
                    raise AuditError("geometry_correction_invalid", f"Поле {path} исправляется более одного раза.")
                room = room_by_id[room_id]
                if not isinstance(room.get("measurements"), dict) or field not in room["measurements"]:
                    raise AuditError("storage_error", f"В geometry отсутствует measurement {path}.")
                room[field] = value
                room["measurements"][field] = dict(measurement)
                changed_paths.add(path)
                applied_targets.append(path)
        else:
            available_element_ids = {
                opening.get("element_id")
                for room_id in selected_room_ids
                for opening in room_by_id[room_id].get(target, [])
                if isinstance(opening, dict) and isinstance(opening.get("element_id"), str)
            }
            selected_element_ids = _correction_selector(
                raw.get("element_ids"),
                available_element_ids,
                f"corrections[{index}].element_ids",
            )
            if not selected_element_ids:
                raise AuditError(
                    "geometry_correction_target_not_found",
                    f"В выбранных помещениях нет элементов target={target!r}.",
                )
            for room_id in sorted(selected_room_ids):
                openings = room_by_id[room_id].get(target)
                if not isinstance(openings, list):
                    raise AuditError("storage_error", f"В geometry повреждена коллекция {room_id}.{target}.")
                for opening in openings:
                    if opening.get("element_id") not in selected_element_ids:
                        continue
                    path = f"{room_id}.{target}.{opening['element_id']}.{field}"
                    if path in changed_paths:
                        raise AuditError("geometry_correction_invalid", f"Поле {path} исправляется более одного раза.")
                    if not isinstance(opening.get("measurements"), dict) or field not in opening["measurements"]:
                        raise AuditError("storage_error", f"В geometry отсутствует measurement {path}.")
                    opening[field] = value
                    opening["measurements"][field] = dict(measurement)
                    changed_paths.add(path)
                    applied_targets.append(path)
            if not applied_targets:
                raise AuditError("geometry_correction_target_not_found", "Правка не выбрала ни одного элемента.")

        normalized.append(
            {
                "target": target,
                "room_ids": raw["room_ids"],
                "element_ids": raw["element_ids"],
                "field": field,
                "input_value": raw.get("value"),
                "normalized_value": value,
                "value": value,
                "applied_targets": applied_targets,
            }
        )

    _remove_stale_geometry_warnings(geometry, normalized)
    geometry = vision.refresh_geometry_derived(geometry)
    review = vision.build_geometry_review(geometry)
    new_revision = geometry_revision + 1
    new_hash = geometry_sha256(geometry)
    created_at = now_iso()
    history_path = output_path(root, "geometry_corrections.json")
    if history_path.exists():
        history = read_json(history_path, code="storage_error")
        if not isinstance(history, list):
            raise AuditError("storage_error", "История исправлений geometry повреждена.")
    else:
        history = []
    history.append(
        {
            "revision_from": geometry_revision,
            "revision_to": new_revision,
            "previous_geometry_sha256": previous_hash,
            "geometry_sha256": new_hash,
            "user_statement": statement,
            "corrections": normalized,
            "task_id": _ctx_value(ctx, "task_id"),
            "client_message_id": _client_message_id(ctx),
            "created_at": created_at,
        }
    )

    # Даже точечная правка меняет основание расчётов. Старые каталог, mapping,
    # фото-выводы и отчёт к новой revision уже не относятся.
    _invalidate_audit(manifest, root)
    atomic_write_json(geometry_path, geometry)
    atomic_write_json(output_path(root, "geometry_review.json"), review)
    atomic_write_json(history_path, history)
    manifest["geometry_revision"] = new_revision
    manifest["skill_version"] = SKILL_VERSION
    manifest["geometry_confirmed"] = False
    manifest["geometry_sha256"] = new_hash
    manifest["geometry_validation_status"] = "validated_with_user_corrections"
    manifest["geometry_correction_count"] = len(history)
    manifest["geometry_corrections_sha256"] = canonical_json_sha256(history)
    manifest["last_geometry_correction_at"] = created_at
    manifest["review_task_id"] = _ctx_value(ctx, "task_id")
    client_message_id = _client_message_id(ctx)
    if client_message_id:
        manifest["review_client_message_id"] = client_message_id
    else:
        manifest.pop("review_client_message_id", None)
    for key in (
        "confirmation_task_id",
        "confirmation_client_message_id",
        "confirmed_geometry_sha256",
        "geometry_confirmed_revision",
        "geometry_confirmed_at",
    ):
        manifest.pop(key, None)
    save_manifest(root, manifest)
    return manifest, {
        "status": "geometry_review_required",
        "geometry_revision": new_revision,
        "corrected_fields": len(changed_paths),
        "confirmation_question": GEOMETRY_CONFIRMATION_NOTICE,
        "next_action": "render_geometry_review",
        "assistant_instruction": GEOMETRY_CORRECTION_ASSISTANT_INSTRUCTION,
    }


def confirm_geometry(
    api: Any,
    ctx: Any,
    job_id: str,
    geometry_revision: Any,
    confirmed: Any,
    corrections: Any = None,
    user_statement: Any = None,
) -> tuple[dict[str, Any], bool, dict[str, Any]]:
    if confirmed is False and corrections is not None:
        manifest, result = revise_geometry(
            api,
            ctx,
            job_id,
            geometry_revision,
            corrections,
            user_statement,
        )
        return manifest, True, result
    if corrections is not None or user_statement is not None:
        raise AuditError(
            "ambiguous_geometry_decision",
            "Нельзя одновременно подтверждать geometry и передавать исправления.",
        )
    root = job_dir(api, job_id)
    manifest = load_manifest(root)
    geometry_path = output_path(root, "geometry.json")
    if not geometry_path.exists():
        raise AuditError("geometry_required", "Сначала сохраните geometry и покажите полный review.")
    if type(geometry_revision) is not int or geometry_revision != manifest.get("geometry_revision"):
        raise AuditError("stale_geometry_revision", "Показанная версия геометрии устарела.")
    if confirmed is not True:
        raise AuditError("explicit_confirmation_required", "confirmed должно быть true после однозначного ответа пользователя.")
    geometry = read_json(geometry_path, code="geometry_required")
    _require_consistent_shared_doors(geometry)
    current_hash = geometry_sha256(geometry)
    if manifest.get("geometry_confirmed") is True:
        # Повтор одного и того же tool-вызова безопасен, пока файл geometry
        # дословно совпадает с уже подтверждённым содержимым.
        if manifest.get("confirmed_geometry_sha256") == current_hash:
            return manifest, False, {
                "status": "geometry_confirmed",
                "geometry_revision": geometry_revision,
                "next_action": "call_mcp_construction_prices__get_supported_works",
                "mcp_tool": PRICE_CATALOG_TOOL,
            }
        raise AuditError("geometry_changed_since_confirmation", "Geometry изменилась после подтверждения.")
    _require_new_geometry_turn(
        manifest,
        ctx,
        code="confirmation_requires_new_turn",
        message="Геометрия должна быть подтверждена отдельным новым сообщением пользователя.",
    )
    current_task_id = _ctx_value(ctx, "task_id")
    current_client_id = _client_message_id(ctx)
    manifest["geometry_confirmed"] = True
    manifest["confirmed_geometry_sha256"] = current_hash
    manifest["geometry_confirmed_revision"] = geometry_revision
    manifest["geometry_confirmed_at"] = now_iso()
    manifest["confirmation_task_id"] = current_task_id
    if current_client_id:
        manifest["confirmation_client_message_id"] = current_client_id
    save_manifest(root, manifest)
    return manifest, True, {
        "status": "geometry_confirmed",
        "geometry_revision": geometry_revision,
        "next_action": "call_mcp_construction_prices__get_supported_works",
        "mcp_tool": PRICE_CATALOG_TOOL,
    }


def _technical_name(value: str) -> str:
    value = " ".join(str(value or "").strip().casefold().replace("ё", "е").split())
    return re.sub(r"[^\w]+", " ", value, flags=re.UNICODE).strip()


def _mapping_text(value: Any, field: str, *, maximum: int = 1000) -> str:
    if not isinstance(value, str) or not value.strip() or len(value) > maximum:
        raise AuditError("invalid_mapping", f"{field} должен быть непустой строкой.")
    return value.strip()


def _confidence(value: Any, field: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(value) or not 0 <= value <= 1:
        raise AuditError("invalid_mapping", f"{field} должен быть finite числом от 0 до 1.")
    return float(value)


def _exact_keys(value: Any, keys: set[str], context: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise AuditError("invalid_mapping", f"{context} должен быть object.")
    unknown = sorted(set(value) - keys)
    missing = sorted(keys - set(value))
    if unknown or missing:
        raise AuditError("invalid_mapping", f"Некорректные поля {context}.", {"unknown": unknown, "missing": missing})
    return value


def _unique_context(estimate: dict[str, Any]) -> tuple[dict[str, str], dict[int, dict[str, Any]]]:
    rooms: dict[str, str] = {}
    works: dict[int, dict[str, Any]] = {}
    for row in estimate["rows"]:
        if row["room"]:
            rooms.setdefault(_technical_name(row["room"]), row["room"])
        if row["work_name"]:
            works[row["source_row"]] = {"name": row["work_name"], "unit": row["unit"]}
    return rooms, works


def _unit_dimension(unit: str) -> str | None:
    normalized = _technical_name(unit)
    if normalized in {_technical_name(item) for item in ("м²", "м2", "кв.м", "м^2")}:
        return "area"
    if normalized in {_technical_name(item) for item in ("м", "пог. м", "п.м")}:
        return "length"
    if normalized in {_technical_name(item) for item in ("шт", "шт.", "ед")}:
        return "count"
    return None


def _units_compatible(left: str, right: str) -> bool:
    if _technical_name(left) == _technical_name(right):
        return True
    left_dimension = _unit_dimension(left)
    right_dimension = _unit_dimension(right)
    return left_dimension is not None and left_dimension == right_dimension


def validate_mapping(
    mapping: Any,
    estimate: dict[str, Any],
    geometry: dict[str, Any],
    delegation_token: str,
    price_catalog: dict[str, Any],
) -> dict[str, Any]:
    required = {
        "schema_version",
        "delegation_token",
        "room_matches",
        "room_unresolved",
        "work_matches",
        "work_unsupported",
        "work_unresolved",
        "price_matches",
        "price_unsupported",
        "price_unresolved",
    }
    root = _exact_keys(mapping, required, "mapping")
    if type(root["schema_version"]) is not int or root["schema_version"] != 3:
        raise AuditError("invalid_mapping", "schema_version должен быть равен 3.")
    if root["delegation_token"] != delegation_token:
        # Старый ответ Mapping может быть корректен по схеме, но относиться к
        # предыдущей geometry revision или каталогу. Токен закрывает этот случай.
        raise AuditError("invalid_mapping", "delegation_token не соответствует текущей подтверждённой geometry revision.")
    for field in required - {"schema_version", "delegation_token"}:
        if not isinstance(root[field], list) or len(root[field]) > 500:
            raise AuditError("invalid_mapping", f"{field} должен быть array не более 500 элементов.")

    estimate_rooms, estimate_works = _unique_context(estimate)
    model_ids = {room["room_id"] for room in geometry["rooms"]}
    canonical_by_technical = {_technical_name(name): name for name in SUPPORTED_WORKS}
    catalog_items = price_catalog.get("items") if isinstance(price_catalog, dict) else None
    if not isinstance(catalog_items, list):
        raise AuditError("price_catalog_invalid", "Сохранённый каталог MCP повреждён.")
    mcp_by_id = {
        item["id"]: item
        for item in catalog_items
        if isinstance(item, dict) and isinstance(item.get("id"), str)
    }
    if len(mcp_by_id) != len(catalog_items):
        raise AuditError("price_catalog_invalid", "Сохранённый каталог MCP содержит некорректные или duplicate id.")
    clean = {
        "schema_version": 3,
        "delegation_token": delegation_token,
        "room_matches": [],
        "room_unresolved": [],
        "work_matches": [],
        "work_unsupported": [],
        "work_unresolved": [],
        "price_matches": [],
        "price_unsupported": [],
        "price_unresolved": [],
    }
    room_seen: set[str] = set()
    for index, raw in enumerate(root["room_matches"]):
        item = _exact_keys(raw, {"estimate_room", "model_room_id", "confidence", "reason"}, f"room_matches[{index}]")
        estimate_room_raw = _mapping_text(item["estimate_room"], "estimate_room", maximum=200)
        estimate_key = _technical_name(estimate_room_raw)
        if estimate_key not in estimate_rooms or estimate_key in room_seen:
            raise AuditError("invalid_mapping", "Неизвестный или duplicate estimate_room.")
        model_id = item["model_room_id"]
        if model_id is None:
            if estimate_key != _technical_name("Весь объект"):
                raise AuditError("invalid_mapping", "model_room_id=null допустим только для «Весь объект».")
        elif not isinstance(model_id, str) or model_id not in model_ids:
            raise AuditError("invalid_mapping", "model_room_id отсутствует в geometry.")
        room_seen.add(estimate_key)
        clean["room_matches"].append(
            {"estimate_room": estimate_rooms[estimate_key], "model_room_id": model_id,
             "confidence": _confidence(item["confidence"], "confidence"), "reason": _mapping_text(item["reason"], "reason")}
        )
    for index, raw in enumerate(root["room_unresolved"]):
        item = _exact_keys(raw, {"estimate_room", "candidate_room_ids", "reason", "requires_human_confirmation"}, f"room_unresolved[{index}]")
        key = _technical_name(_mapping_text(item["estimate_room"], "estimate_room", maximum=200))
        candidates = item["candidate_room_ids"]
        if key not in estimate_rooms or key in room_seen or not isinstance(candidates, list) or len(candidates) != len(set(candidates)) or any(candidate not in model_ids for candidate in candidates) or type(item["requires_human_confirmation"]) is not bool:
            raise AuditError("invalid_mapping", "Некорректная unresolved room mapping.")
        room_seen.add(key)
        clean["room_unresolved"].append(
            {"estimate_room": estimate_rooms[key], "candidate_room_ids": candidates,
             "reason": _mapping_text(item["reason"], "reason"), "requires_human_confirmation": item["requires_human_confirmation"]}
        )
    if room_seen != set(estimate_rooms):
        raise AuditError("invalid_mapping", "Mapping не покрывает все помещения.", {"missing": sorted(set(estimate_rooms) - room_seen)})

    # Покрытие считаем по source_row, а не по названию: одинаковая работа в двух
    # строках может иметь разные единицы и разные решения Mapping.
    work_seen: set[int] = set()
    for index, raw in enumerate(root["work_matches"]):
        item = _exact_keys(raw, {"source_row", "estimate_work", "canonical_work", "confidence", "reason"}, f"work_matches[{index}]")
        source_row = item["source_row"]
        estimate_key = _technical_name(_mapping_text(item["estimate_work"], "estimate_work", maximum=200))
        canonical_key = _technical_name(_mapping_text(item["canonical_work"], "canonical_work", maximum=200))
        if type(source_row) is not int or source_row not in estimate_works or source_row in work_seen or canonical_key not in canonical_by_technical:
            raise AuditError("invalid_mapping", "Неизвестная или duplicate work mapping.")
        estimate_work = estimate_works[source_row]
        if estimate_key != _technical_name(estimate_work["name"]):
            raise AuditError("invalid_mapping", "estimate_work не соответствует source_row.")
        canonical = canonical_by_technical[canonical_key]
        expected_dimension = SUPPORTED_WORKS[canonical][2]
        recognized_dimension = _unit_dimension(estimate_work["unit"])
        if recognized_dimension is not None and recognized_dimension != expected_dimension:
            raise AuditError("incompatible_unit", "Единица работы несовместима с canonical work.", {"source_row": source_row, "estimate_work": estimate_work["name"], "canonical_work": canonical})
        work_seen.add(source_row)
        clean["work_matches"].append(
            {"source_row": source_row, "estimate_work": estimate_work["name"], "canonical_work": canonical,
             "confidence": _confidence(item["confidence"], "confidence"), "reason": _mapping_text(item["reason"], "reason")}
        )
    for field, keys in (
        ("work_unsupported", {"source_row", "estimate_work", "reason"}),
        ("work_unresolved", {"source_row", "estimate_work", "candidate_canonical_works", "reason", "requires_human_confirmation"}),
    ):
        for index, raw in enumerate(root[field]):
            item = _exact_keys(raw, keys, f"{field}[{index}]")
            source_row = item["source_row"]
            estimate_key = _technical_name(_mapping_text(item["estimate_work"], "estimate_work", maximum=200))
            if type(source_row) is not int or source_row not in estimate_works or source_row in work_seen:
                raise AuditError("invalid_mapping", "Неизвестная или overlapping estimate_work.")
            estimate_work = estimate_works[source_row]
            if estimate_key != _technical_name(estimate_work["name"]):
                raise AuditError("invalid_mapping", "estimate_work не соответствует source_row.")
            work_seen.add(source_row)
            if field == "work_unsupported":
                clean[field].append({"source_row": source_row, "estimate_work": estimate_work["name"], "reason": _mapping_text(item["reason"], "reason")})
            else:
                candidates = item["candidate_canonical_works"]
                if not isinstance(candidates, list) or len(candidates) != len(set(candidates)) or any(_technical_name(candidate) not in canonical_by_technical for candidate in candidates) or type(item["requires_human_confirmation"]) is not bool:
                    raise AuditError("invalid_mapping", "Некорректная unresolved work mapping.")
                clean[field].append(
                    {"source_row": source_row, "estimate_work": estimate_work["name"],
                     "candidate_canonical_works": [canonical_by_technical[_technical_name(candidate)] for candidate in candidates],
                     "reason": _mapping_text(item["reason"], "reason"), "requires_human_confirmation": item["requires_human_confirmation"]}
                )
    if work_seen != set(estimate_works):
        raise AuditError("invalid_mapping", "Mapping не покрывает все работы.", {"missing": sorted(set(estimate_works) - work_seen)})

    price_seen: set[int] = set()
    for index, raw in enumerate(root["price_matches"]):
        item = _exact_keys(
            raw,
            {"source_row", "estimate_work", "mcp_work_id", "confidence", "reason"},
            f"price_matches[{index}]",
        )
        source_row = item["source_row"]
        estimate_key = _technical_name(_mapping_text(item["estimate_work"], "estimate_work", maximum=200))
        mcp_work_id = _mapping_text(item["mcp_work_id"], "mcp_work_id", maximum=200)
        if type(source_row) is not int or source_row not in estimate_works or source_row in price_seen:
            raise AuditError("invalid_mapping", "Неизвестная или duplicate price mapping.")
        estimate_work = estimate_works[source_row]
        if estimate_key != _technical_name(estimate_work["name"]):
            raise AuditError("invalid_mapping", "estimate_work не соответствует source_row.")
        catalog_item = mcp_by_id.get(mcp_work_id)
        if catalog_item is None:
            raise AuditError(
                "invalid_mapping",
                "mcp_work_id отсутствует в сохранённом каталоге MCP.",
                {"source_row": source_row, "mcp_work_id": mcp_work_id},
            )
        if not _units_compatible(estimate_work["unit"], str(catalog_item.get("unit") or "")):
            raise AuditError(
                "incompatible_unit",
                "Единица сметы несовместима с единицей позиции MCP.",
                {
                    "source_row": source_row,
                    "estimate_work": estimate_work["name"],
                    "estimate_unit": estimate_work["unit"],
                    "mcp_work_id": mcp_work_id,
                    "mcp_unit": catalog_item.get("unit"),
                },
            )
        price_seen.add(source_row)
        clean["price_matches"].append(
            {
                "source_row": source_row,
                "estimate_work": estimate_work["name"],
                "mcp_work_id": mcp_work_id,
                "confidence": _confidence(item["confidence"], "confidence"),
                "reason": _mapping_text(item["reason"], "reason"),
            }
        )

    for field, keys in (
        ("price_unsupported", {"source_row", "estimate_work", "reason"}),
        (
            "price_unresolved",
            {
                "source_row",
                "estimate_work",
                "candidate_mcp_work_ids",
                "reason",
                "requires_human_confirmation",
            },
        ),
    ):
        for index, raw in enumerate(root[field]):
            item = _exact_keys(raw, keys, f"{field}[{index}]")
            source_row = item["source_row"]
            estimate_key = _technical_name(_mapping_text(item["estimate_work"], "estimate_work", maximum=200))
            if type(source_row) is not int or source_row not in estimate_works or source_row in price_seen:
                raise AuditError("invalid_mapping", "Неизвестная или overlapping price mapping.")
            estimate_work = estimate_works[source_row]
            if estimate_key != _technical_name(estimate_work["name"]):
                raise AuditError("invalid_mapping", "estimate_work не соответствует source_row.")
            price_seen.add(source_row)
            if field == "price_unsupported":
                clean[field].append(
                    {
                        "source_row": source_row,
                        "estimate_work": estimate_work["name"],
                        "reason": _mapping_text(item["reason"], "reason"),
                    }
                )
            else:
                candidates = item["candidate_mcp_work_ids"]
                if (
                    not isinstance(candidates, list)
                    or any(not isinstance(candidate, str) for candidate in candidates)
                    or len(candidates) != len(set(candidates))
                    or any(candidate not in mcp_by_id for candidate in candidates)
                    or type(item["requires_human_confirmation"]) is not bool
                ):
                    raise AuditError("invalid_mapping", "Некорректная unresolved price mapping.")
                clean[field].append(
                    {
                        "source_row": source_row,
                        "estimate_work": estimate_work["name"],
                        "candidate_mcp_work_ids": candidates,
                        "reason": _mapping_text(item["reason"], "reason"),
                        "requires_human_confirmation": item["requires_human_confirmation"],
                    }
                )
    if price_seen != set(estimate_works):
        raise AuditError(
            "invalid_mapping",
            "Ценовой Mapping не покрывает все работы.",
            {"missing": sorted(set(estimate_works) - price_seen)},
        )
    return clean


def _sum_or_none(values: list[Decimal | None]) -> Decimal | None:
    if any(value is None for value in values):
        return None
    return sum((value for value in values if value is not None), Decimal(0))


def _opening_area(openings: list[dict[str, Any]]) -> Decimal | None:
    values: list[Decimal | None] = []
    for opening in openings:
        width = Decimal(opening["width_m"]) if opening["width_m"] is not None else None
        height = Decimal(opening["height_m"]) if opening["height_m"] is not None else None
        values.append(width * height if width is not None and height is not None else None)
    return _sum_or_none(values)


def _unique_opening_count(rooms: list[dict[str, Any]], collection: str) -> Decimal:
    return Decimal(len({opening["element_id"] for room in rooms for opening in room[collection]}))


def calculate_quantities(geometry: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]]]:
    metric_names = (
        "floor_area_m2",
        "ceiling_area_m2",
        "gross_wall_area_m2",
        "doors_area_m2",
        "windows_area_m2",
        "openings_area_m2",
        "net_wall_area_m2",
        "baseboard_length_m",
        "door_count",
        "window_count",
    )
    rooms_out: list[dict[str, Any]] = []
    trace_entries: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    # Неизвестный размер не заменяем нулём. None должен дойти только до зависимых
    # метрик, чтобы остальные строки аудита всё равно можно было проверить.
    raw_by_room: dict[str, dict[str, Decimal | None]] = {}
    for room in geometry["rooms"]:
        floor = Decimal(room["floor_area_m2"]) if room["floor_area_m2"] is not None else None
        perimeter = Decimal(room["perimeter_m"]) if room["perimeter_m"] is not None else None
        height = Decimal(room["height_m"]) if room["height_m"] is not None else None
        doors_area = _opening_area(room["doors"])
        windows_area = _opening_area(room["windows"])
        openings_area = doors_area + windows_area if doors_area is not None and windows_area is not None else None
        gross = perimeter * height if perimeter is not None and height is not None else None
        net = gross - openings_area if gross is not None and openings_area is not None else None
        door_widths = [Decimal(item["width_m"]) if item["width_m"] is not None else None for item in room["doors"]]
        door_width_total = _sum_or_none(door_widths)
        baseboard = perimeter - door_width_total if perimeter is not None and door_width_total is not None else None
        if net is not None and net < 0:
            warnings.append({"code": "geometry_conflict", "room_id": room["room_id"], "message": "Площадь проёмов превышает gross wall area."})
            net = None
        if baseboard is not None and baseboard < 0:
            warnings.append({"code": "geometry_conflict", "room_id": room["room_id"], "message": "Ширина дверей превышает периметр."})
            baseboard = None
        raw = {
            "floor_area_m2": floor,
            "ceiling_area_m2": floor,
            "gross_wall_area_m2": gross,
            "doors_area_m2": doors_area,
            "windows_area_m2": windows_area,
            "openings_area_m2": openings_area,
            "net_wall_area_m2": net,
            "baseboard_length_m": baseboard,
            "door_count": Decimal(len(room["doors"])),
            "window_count": Decimal(len(room["windows"])),
        }
        raw_by_room[room["room_id"]] = raw
        metrics: dict[str, str | None] = {}
        for metric in metric_names:
            integer = metric in {"door_count", "window_count"}
            metrics[metric] = display_decimal(raw[metric], integer=integer)
            trace_id = f"room:{room['room_id']}:{metric}"
            inputs: dict[str, Any]
            formula: str
            if metric == "floor_area_m2":
                formula, inputs = "explicit floor_area_m2", {"floor_area_m2": room["floor_area_m2"]}
            elif metric == "ceiling_area_m2":
                formula, inputs = "floor_area_m2", {"floor_area_m2": room["floor_area_m2"]}
            elif metric == "gross_wall_area_m2":
                formula, inputs = "perimeter_m * height_m", {"perimeter_m": room["perimeter_m"], "height_m": room["height_m"]}
            elif metric in {"doors_area_m2", "windows_area_m2"}:
                collection = "doors" if metric.startswith("doors") else "windows"
                formula = f"sum({collection}.width_m * {collection}.height_m)"
                inputs = {collection: [{"element_id": item["element_id"], "width_m": item["width_m"], "height_m": item["height_m"]} for item in room[collection]]}
            elif metric == "openings_area_m2":
                formula, inputs = "doors_area_m2 + windows_area_m2", {"doors_area_m2": decimal_text(doors_area), "windows_area_m2": decimal_text(windows_area)}
            elif metric == "net_wall_area_m2":
                formula, inputs = "gross_wall_area_m2 - openings_area_m2", {"gross_wall_area_m2": decimal_text(gross), "openings_area_m2": decimal_text(openings_area)}
            elif metric == "baseboard_length_m":
                formula, inputs = "perimeter_m - sum(unique room door widths)", {"perimeter_m": room["perimeter_m"], "doors": [{"element_id": item["element_id"], "width_m": item["width_m"]} for item in room["doors"]]}
            elif metric == "door_count":
                formula, inputs = "len(room doors)", {"element_ids": [item["element_id"] for item in room["doors"]]}
            else:
                formula, inputs = "len(room windows)", {"element_ids": [item["element_id"] for item in room["windows"]]}
            entry_warnings = [] if raw[metric] is not None else ["missing_geometry"]
            if entry_warnings:
                warnings.append({"code": "missing_geometry", "room_id": room["room_id"], "metric": metric, "message": f"Недостаточно геометрии для {metric}."})
            trace_entries.append(
                {"trace_id": trace_id, "room_id": room["room_id"], "room_name": room["name"],
                 "metric": metric, "formula": formula, "inputs": inputs,
                 "raw_result": decimal_text(raw[metric]), "rounded_result": metrics[metric], "warnings": entry_warnings}
            )
        rooms_out.append({"room_id": room["room_id"], "room_name": room["name"], "metrics": metrics})

    # Общий межкомнатный проём участвует в стенах обеих комнат, но установкой
    # считается один раз на объект — поэтому counts дедуплицируются по element_id.
    totals_raw: dict[str, Decimal | None] = {}
    for metric in metric_names:
        if metric == "door_count":
            totals_raw[metric] = _unique_opening_count(geometry["rooms"], "doors")
        elif metric == "window_count":
            totals_raw[metric] = _unique_opening_count(geometry["rooms"], "windows")
        else:
            totals_raw[metric] = _sum_or_none([raw_by_room[room["room_id"]][metric] for room in geometry["rooms"]])
        trace_entries.append(
            {"trace_id": f"object_total:{metric}", "room_id": None, "room_name": "Весь объект",
             "metric": metric, "formula": "deduplicate by element_id" if metric in {"door_count", "window_count"} else "sum(room metric)",
             "inputs": {
                 "room_values": [{"room_id": room["room_id"], "value": decimal_text(raw_by_room[room["room_id"]][metric])} for room in geometry["rooms"]],
                 **({"element_ids": sorted({item["element_id"] for room in geometry["rooms"] for item in room["doors" if metric == "door_count" else "windows"]})} if metric in {"door_count", "window_count"} else {}),
             },
             "raw_result": decimal_text(totals_raw[metric]),
             "rounded_result": display_decimal(totals_raw[metric], integer=metric in {"door_count", "window_count"}),
             "warnings": [] if totals_raw[metric] is not None else ["missing_geometry"]}
        )
    quantities = {
        "schema_version": 1,
        "rooms": rooms_out,
        "object_totals": {metric: display_decimal(value, integer=metric in {"door_count", "window_count"}) for metric, value in totals_raw.items()},
    }
    trace = {"schema_version": 1, "entries": trace_entries}
    deduped_warnings: list[dict[str, Any]] = []
    seen_warning: set[str] = set()
    for warning in warnings:
        key = json.dumps(warning, ensure_ascii=False, sort_keys=True)
        if key not in seen_warning:
            seen_warning.add(key)
            deduped_warnings.append(warning)
    return quantities, trace, deduped_warnings


def _generated_estimate_catalog_item(
    work_name: str,
    price_catalog: dict[str, Any],
) -> dict[str, Any] | None:
    """Return only one exact canonical-name and compatible-unit MCP item."""
    allowed_units = {
        _technical_name(unit) for unit in SUPPORTED_WORKS[work_name][1]
    }
    candidates = [
        item
        for item in price_catalog.get("items", [])
        if _technical_name(item.get("name", "")) == _technical_name(work_name)
        and _technical_name(item.get("unit", "")) in allowed_units
    ]
    return candidates[0] if len(candidates) == 1 else None


def _excel_text(value: Any) -> str:
    text = str(value or "")
    # Названия приходят от пользователя и MCP. Апостроф не даёт Excel
    # выполнить строку, начинающуюся как формула, при открытии сметы.
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def _excel_number(value: Decimal) -> int | float:
    return int(value) if value == value.to_integral_value() else float(value)


def generate_estimate(
    api: Any,
    job_id: str,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Create a simple preliminary XLSX from confirmed geometry and saved MCP prices."""
    root = job_dir(api, job_id)
    manifest = load_manifest(root)
    if not manifest.get("geometry_confirmed"):
        raise AuditError(
            "geometry_confirmation_required",
            "Сначала подтвердите geometry отдельным новым сообщением пользователя.",
        )
    geometry = read_json(output_path(root, "geometry.json"), code="geometry_required")
    geometry_fingerprint = geometry_sha256(geometry)
    if geometry_fingerprint != manifest.get("confirmed_geometry_sha256"):
        raise AuditError(
            "geometry_changed_since_confirmation",
            "Подтверждённый hash geometry не совпадает с текущим.",
        )
    if manifest.get("geometry_confirmed_revision") != manifest.get("geometry_revision"):
        raise AuditError(
            "stale_geometry_confirmation",
            "Подтверждение относится не к текущей geometry revision.",
        )
    derived_issues = vision.derived_state_issues(geometry)
    if derived_issues:
        raise AuditError(
            "geometry_derived_state_incomplete",
            "Для части помещений есть исходные размеры, но не рассчитаны производные значения.",
            {"incomplete_fields": derived_issues},
        )

    price_catalog = read_json(
        output_path(root, "price_catalog.json"), code="price_catalog_required"
    )
    catalog_fingerprint = canonical_json_sha256(price_catalog)
    if catalog_fingerprint != manifest.get("price_catalog_sha256"):
        raise AuditError(
            "price_catalog_changed",
            "Сохранённый каталог MCP изменился после validation.",
        )

    quantities, _, quantity_warnings = calculate_quantities(geometry)
    catalog_by_work = {
        work_name: _generated_estimate_catalog_item(work_name, price_catalog)
        for work_name in SUPPORTED_WORKS
    }
    notes = {
        "Грунтовка стен": "Чистая площадь стен без проёмов",
        "Окраска стен": "Чистая площадь стен без проёмов",
        "Устройство пола": "По площади помещения",
        "Отделка потолка": "По площади помещения",
        "Монтаж плинтуса": "Периметр минус ширина дверных проёмов",
        "Установка окон": "По количеству окон помещения",
        "Установка дверей": "Уникальные двери по всему объекту",
    }
    generated_rows: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for room in quantities["rooms"]:
        for work_name, (metric, _, _) in SUPPORTED_WORKS.items():
            if work_name == DOOR_INSTALLATION_WORK:
                continue
            catalog_item = catalog_by_work[work_name]
            quantity_text = room["metrics"].get(metric)
            if catalog_item is None:
                skipped.append(
                    {"room": room["room_name"], "work": work_name, "reason": "mcp_exact_match_missing"}
                )
                continue
            if quantity_text is None:
                skipped.append(
                    {"room": room["room_name"], "work": work_name, "reason": "missing_geometry"}
                )
                continue
            quantity = Decimal(quantity_text)
            if quantity <= 0:
                continue
            price = Decimal(catalog_item["price"])
            generated_rows.append(
                {
                    "room": room["room_name"],
                    "work_name": work_name,
                    "unit": catalog_item["unit"],
                    "quantity": decimal_text(quantity),
                    "price": decimal_text(price),
                    "total": decimal_text(quantity * price),
                    "note": notes[work_name],
                    "mcp_work_id": catalog_item["id"],
                }
            )

    door_item = catalog_by_work[DOOR_INSTALLATION_WORK]
    door_quantity_text = quantities["object_totals"].get("door_count")
    if door_item is None:
        skipped.append(
            {"room": "Весь объект", "work": DOOR_INSTALLATION_WORK, "reason": "mcp_exact_match_missing"}
        )
    elif door_quantity_text is None:
        skipped.append(
            {"room": "Весь объект", "work": DOOR_INSTALLATION_WORK, "reason": "missing_geometry"}
        )
    else:
        door_quantity = Decimal(door_quantity_text)
        if door_quantity > 0:
            door_price = Decimal(door_item["price"])
            generated_rows.append(
                {
                    "room": "Весь объект",
                    "work_name": DOOR_INSTALLATION_WORK,
                    "unit": door_item["unit"],
                    "quantity": decimal_text(door_quantity),
                    "price": decimal_text(door_price),
                    "total": decimal_text(door_quantity * door_price),
                    "note": notes[DOOR_INSTALLATION_WORK],
                    "mcp_work_id": door_item["id"],
                }
            )

    if not generated_rows:
        raise AuditError(
            "generated_estimate_empty",
            "Не удалось сформировать ни одной строки сметы по текущей geometry и MCP-каталогу.",
            {"skipped": skipped[:100]},
        )

    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise AuditError("dependency_missing", "Для XLSX требуется openpyxl.") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Смета"
    sheet.append(list(REQUIRED_HEADERS))
    sheet.row_dimensions[1].height = 42.75
    widths = {"A": 7, "B": 18, "C": 26, "D": 12, "E": 14, "F": 18, "G": 16, "H": 34}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for cell in sheet[1]:
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for position, item in enumerate(generated_rows, start=1):
        excel_row = position + 1
        sheet.append(
            [
                position,
                _excel_text(item["room"]),
                _excel_text(item["work_name"]),
                _excel_text(item["unit"]),
                _excel_number(Decimal(item["quantity"])),
                _excel_number(Decimal(item["price"])),
                f"=E{excel_row}*F{excel_row}",
                _excel_text(item["note"]),
            ]
        )
        sheet.cell(excel_row, 6).comment = Comment(
            f"Источник цены: {PRICE_CATALOG_TOOL}; MCP work id: {item['mcp_work_id']}",
            "Construction Audit MVP",
        )
        for column in range(1, 9):
            sheet.cell(excel_row, column).font = Font(name="Arial", size=11)
        for column in (5, 6, 7):
            sheet.cell(excel_row, column).number_format = "#,##0.##"

    table = Table(displayName="GeneratedEstimateTable", ref=f"A1:H{len(generated_rows) + 1}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    destination = output_path(root, "generated_estimate.xlsx")
    temporary: str | None = None
    try:
        with tempfile.NamedTemporaryFile(
            dir=destination.parent, prefix=".generated_estimate.", suffix=".xlsx", delete=False
        ) as handle:
            temporary = handle.name
        workbook.save(temporary)
        _validate_xlsx(Path(temporary))
        os.replace(temporary, destination)
        temporary = None
    except AuditError:
        raise
    except Exception as exc:
        raise AuditError("generated_estimate_failed", "Не удалось создать XLSX-смету.") from exc
    finally:
        workbook.close()
        if temporary:
            Path(temporary).unlink(missing_ok=True)

    artifact_sha256 = sha256_file(destination)
    metadata = {
        "schema_version": 1,
        "geometry_revision": manifest["geometry_revision"],
        "geometry_sha256": geometry_fingerprint,
        "price_catalog_sha256": catalog_fingerprint,
        "rows": generated_rows,
        "skipped": skipped,
        "quantity_warnings": quantity_warnings,
        "xlsx_sha256": artifact_sha256,
    }
    atomic_write_json(output_path(root, "generated_estimate.json"), metadata)
    manifest["generated_estimate_sha256"] = artifact_sha256
    manifest["generated_estimate_geometry_sha256"] = geometry_fingerprint
    manifest["generated_estimate_price_catalog_sha256"] = catalog_fingerprint
    manifest["generated_estimate_rows_count"] = len(generated_rows)
    manifest["generated_estimate_created_at"] = now_iso()
    manifest["skill_version"] = SKILL_VERSION
    save_manifest(root, manifest)
    return manifest, {
        "status": "estimate_generated",
        "rows_count": len(generated_rows),
        "skipped_count": len(skipped),
        "estimate_artifact": {
            "path": str(destination),
            "size_bytes": destination.stat().st_size,
            "sha256": artifact_sha256,
        },
        "next_action": "display_estimate_artifact",
    }


def _safe_summary(kind: str, source_row: int, estimated: str | None, control: str | None, unit: str) -> str:
    if kind in {"quantity_overstatement", "quantity_understatement"}:
        direction = "выше" if kind == "quantity_overstatement" else "ниже"
        return (
            f"По строке {source_row} выявлено предварительное расхождение: значение сметы {estimated or '—'} {unit} "
            f"{direction} контрольного расчётного значения {control or '—'} {unit}; строка требует проверки специалистом."
        )
    return f"По строке {source_row} выявлено предварительное расхождение типа {kind}; строка требует проверки специалистом."


def _finding(
    kind: str,
    row: dict[str, Any],
    *,
    canonical_room_id: str | None,
    canonical_room_name: str,
    canonical_work: str,
    estimated: Decimal | None,
    control: Decimal | None,
    unit: str,
    trace_refs: list[str],
    source_rows: list[int] | None = None,
    severity: str = "warning",
) -> dict[str, Any]:
    deviation_absolute = abs(estimated - control) if estimated is not None and control is not None else None
    deviation_percent = (
        deviation_absolute / abs(control) * Decimal(100)
        if deviation_absolute is not None and control not in (None, Decimal(0))
        else None
    )
    result = {
        "type": kind,
        "source_row": row["source_row"],
        "source_position": row.get("position"),
        "source_rows": source_rows or [row["source_row"]],
        "original_room_name": row["room"],
        "canonical_room_id": canonical_room_id,
        "canonical_room_name": canonical_room_name,
        "original_work_name": row["work_name"],
        "canonical_work_name": canonical_work,
        "estimated_value": decimal_text(estimated),
        "control_value": decimal_text(control),
        "unit": unit,
        "deviation_absolute": decimal_text(deviation_absolute),
        "deviation_percent": decimal_text(deviation_percent),
        "calculation_trace_refs": trace_refs,
        "severity": severity,
        "status": "preliminary",
    }
    result["safe_summary"] = _safe_summary(kind, row["source_row"], result["estimated_value"], result["control_value"], unit)
    return result


def run_checks(
    estimate: dict[str, Any],
    geometry: dict[str, Any],
    mapping: dict[str, Any],
    quantities: dict[str, Any],
    trace: dict[str, Any],
    tolerance: Decimal,
    quantity_checks: list[dict[str, Any]] | None = None,
) -> tuple[
    list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]],
    list[dict[str, Any]],
]:
    findings: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = list(estimate.get("warnings", []))
    checked_rows: list[dict[str, Any]] = []
    not_checked_rows: list[dict[str, Any]] = []
    if quantity_checks is None:
        quantity_checks = []
    room_map = {_technical_name(item["estimate_room"]): item["model_room_id"] for item in mapping["room_matches"]}
    work_map = {item["source_row"]: item["canonical_work"] for item in mapping["work_matches"]}
    work_mapping_by_row = {item["source_row"]: item for item in mapping["work_matches"]}
    unsupported = {item["source_row"]: item for item in mapping["work_unsupported"]}
    geometry_rooms = {room["room_id"]: room for room in geometry["rooms"]}
    quantity_rooms = {room["room_id"]: room for room in quantities["rooms"]}
    trace_refs: dict[tuple[str, str | None], list[str]] = {}
    for entry in trace["entries"]:
        trace_refs.setdefault((entry["metric"], entry["room_id"]), []).append(entry["trace_id"])
    raw_control_by_metric_room = {
        (entry.get("metric"), entry.get("room_id")): entry.get("raw_result")
        for entry in trace["entries"]
        if "raw_result" in entry
    }

    duplicate_seen: dict[tuple[Any, ...], int] = {}
    groups: dict[tuple[str | None, str, str], list[dict[str, Any]]] = {}
    for row in estimate["rows"]:
        for issue in row["issues"]:
            if issue["type"] in {"invalid_quantity", "invalid_price", "invalid_total", "arithmetic_mismatch"}:
                quantity_value = _decimal(row.get("quantity"))
                price_value = _decimal(row.get("price"))
                total_value = _decimal(row.get("total"))
                if issue["type"] == "invalid_quantity":
                    issue_estimated, issue_control, issue_unit = quantity_value, None, row["unit"]
                elif issue["type"] == "invalid_price":
                    issue_estimated, issue_control, issue_unit = price_value, None, ""
                elif issue["type"] == "invalid_total":
                    issue_estimated = total_value
                    issue_control = quantity_value * price_value if quantity_value is not None and price_value is not None else None
                    issue_unit = ""
                else:
                    issue_estimated = total_value
                    issue_control = quantity_value * price_value if quantity_value is not None and price_value is not None else None
                    issue_unit = ""
                findings.append(
                    _finding(issue["type"], row, canonical_room_id=None, canonical_room_name="—",
                             canonical_work="—", estimated=issue_estimated, control=issue_control,
                             unit=issue_unit, trace_refs=[], severity="high" if issue["type"] == "invalid_quantity" else "warning")
                )
        room_key = _technical_name(row["room"])
        source_row = row["source_row"]
        if source_row in unsupported:
            warnings.append({"code": "unsupported_work", "source_row": source_row, "work_name": row["work_name"], "message": unsupported[source_row]["reason"]})
            not_checked_rows.append({"source_row": row["source_row"], "status": "not_checked", "reason": "unsupported_work"})
            continue
        if row["quantity"] is None:
            not_checked_rows.append({"source_row": row["source_row"], "status": "not_checked", "reason": "invalid_quantity"})
            continue
        if room_key not in room_map or source_row not in work_map:
            code = "unmapped_room" if room_key not in room_map else "unmapped_work"
            warnings.append({"code": code, "source_row": row["source_row"], "message": "Mapping отсутствует."})
            not_checked_rows.append({"source_row": row["source_row"], "status": "not_checked", "reason": code})
            continue
        room_id = room_map[room_key]
        canonical_work = work_map[source_row]
        mapping_reason = str(work_mapping_by_row.get(source_row, {}).get("reason") or "")
        if any(marker in mapping_reason.casefold() for marker in ("опечат", "typo")):
            warnings.append({
                "code": "source_work_name_typo",
                "level": "info",
                "source_row": source_row,
                "position": row.get("position"),
                "message": f"Название «{row['work_name']}» нормализовано как «{canonical_work}»; вероятна опечатка.",
            })
        duplicate_key = (room_id, canonical_work, _technical_name(row["unit"]), row["quantity"], row["price"], row["total"])
        if duplicate_key in duplicate_seen:
            findings.append(
                _finding("exact_duplicate", row, canonical_room_id=room_id,
                         canonical_room_name="Весь объект" if room_id is None else geometry_rooms[room_id]["name"],
                         canonical_work=canonical_work, estimated=Decimal(row["quantity"]), control=None,
                         unit=row["unit"], trace_refs=[], severity="high")
            )
        else:
            duplicate_seen[duplicate_key] = row["source_row"]
        group_room_id = None if canonical_work == DOOR_INSTALLATION_WORK else room_id
        group_unit = row["unit"]
        if canonical_work == DOOR_INSTALLATION_WORK:
            allowed_units = SUPPORTED_WORKS[canonical_work][1]
            if _technical_name(group_unit) in {_technical_name(item) for item in allowed_units}:
                group_unit = allowed_units[0]
        groups.setdefault((group_room_id, canonical_work, group_unit), []).append(row)

    for (room_id, canonical_work, unit), rows in groups.items():
        metric, allowed_units, _ = SUPPORTED_WORKS[canonical_work]
        estimated = sum((Decimal(row["quantity"]) for row in rows), Decimal(0))
        is_door_object_check = canonical_work == DOOR_INSTALLATION_WORK
        room_name = "Весь объект" if room_id is None else geometry_rooms[room_id]["name"]
        allowed_normalized = {_technical_name(item) for item in allowed_units}
        if _technical_name(unit) not in allowed_normalized:
            findings.append(
                _finding("unit_mismatch", rows[0], canonical_room_id=room_id, canonical_room_name=room_name,
                         canonical_work=canonical_work, estimated=estimated, control=None, unit=unit,
                         trace_refs=trace_refs.get((metric, room_id), []), severity="high")
            )
            for row in rows:
                not_checked_rows.append({"source_row": row["source_row"], "status": "not_checked", "reason": "unit_mismatch"})
            continue
        control_display_text = quantities["object_totals"][metric] if room_id is None else quantity_rooms[room_id]["metrics"][metric]
        control_raw_text = raw_control_by_metric_room.get((metric, room_id))
        if control_raw_text is None:
            for row in rows:
                warnings.append({"code": "missing_geometry", "source_row": row["source_row"], "metric": metric, "message": "Недостаточно геометрии для контрольного расчёта."})
                not_checked_rows.append({"source_row": row["source_row"], "status": "not_checked", "reason": "missing_geometry"})
            continue
        control = Decimal(control_raw_text)
        deviation = abs(estimated - control) / abs(control) * Decimal(100) if control else (Decimal(0) if estimated == 0 else Decimal("Infinity"))
        threshold_exceeded = not deviation.is_finite() or deviation > tolerance
        if estimated == control:
            quantity_status = "exact_match"
        elif threshold_exceeded:
            quantity_status = "deviation_found"
        else:
            quantity_status = "below_threshold"
        quantity_trace_id = "quantity_check:rows:" + "-".join(str(row["source_row"]) for row in rows)
        trace["entries"].append({
            "trace_id": quantity_trace_id,
            "source_rows": [row["source_row"] for row in rows],
            "room_id": room_id,
            "metric": "quantity_threshold_check",
            "formula": "abs(estimate_quantity - control_quantity) / abs(control_quantity) * 100 > tolerance_percent",
            "inputs": {
                "estimate_quantity": decimal_text(estimated),
                "control_quantity_raw": decimal_text(control),
                "control_quantity_display": control_display_text,
                "tolerance_percent": decimal_text(tolerance),
                "comparison_operator": ">",
            },
            "results": {
                "deviation_signed": decimal_text(estimated - control),
                "deviation_absolute": decimal_text(abs(estimated - control)),
                "deviation_percent_raw": decimal_text(deviation) if deviation.is_finite() else None,
                "threshold_exceeded": threshold_exceeded,
                "status": quantity_status,
            },
            "warnings": ["zero_control_quantity"] if control == 0 and estimated != 0 else [],
        })
        for row in rows:
            quantity_checks.append({
                "source_row": row["source_row"],
                "source_position": row.get("position"),
                "source_rows": [item["source_row"] for item in rows],
                "work_name": row["work_name"],
                "canonical_work": canonical_work,
                "estimated_value": decimal_text(estimated),
                "control_value": decimal_text(control),
                "control_display_value": control_display_text,
                "deviation_signed": decimal_text(estimated - control),
                "deviation_absolute": decimal_text(abs(estimated - control)),
                "deviation_percent": decimal_text(deviation) if deviation.is_finite() else None,
                "tolerance_percent": decimal_text(tolerance),
                "comparison_operator": ">",
                "threshold_exceeded": threshold_exceeded,
                "status": quantity_status,
                "trace_ref": quantity_trace_id,
            })
        if threshold_exceeded:
            kind = "quantity_overstatement" if estimated > control else "quantity_understatement"
            severity = "high" if not deviation.is_finite() or deviation > Decimal(20) else "warning"
            finding = _finding(
                kind,
                rows[0],
                canonical_room_id=room_id,
                canonical_room_name=room_name,
                canonical_work=canonical_work,
                estimated=estimated,
                control=control,
                unit=unit,
                trace_refs=[*trace_refs.get((metric, room_id), []), quantity_trace_id],
                source_rows=[row["source_row"] for row in rows],
                severity=severity,
            )
            if is_door_object_check:
                finding["quantity_check_scope"] = DOOR_QUANTITY_SCOPE
                finding["allocation_status"] = ESTIMATE_DECLARED_ALLOCATION
                finding["original_room_name"] = "Все помещения (по смете)"
                source_rows_text = ", ".join(str(row["source_row"]) for row in rows)
                finding["safe_summary"] = (
                    f"По строкам {source_rows_text} суммарное количество установок дверей по смете "
                    f"составляет {finding['estimated_value'] or '—'} {unit}, а уникальное количество "
                    f"дверей на плане — {finding['control_value'] or '—'} {unit}; распределение по "
                    "помещениям принято из сметы и отдельно не проверялось."
                )
            findings.append(finding)
        for row in rows:
            checked = {
                "source_row": row["source_row"],
                "source_position": row.get("position"),
                "status": "checked",
                "work_name": row["work_name"],
                "control_value": decimal_text(control),
                "quantity_result": quantity_status,
                "deviation_percent": decimal_text(deviation) if deviation.is_finite() else None,
            }
            if is_door_object_check:
                checked.update(
                    {
                        "quantity_check_scope": DOOR_QUANTITY_SCOPE,
                        "allocation_status": ESTIMATE_DECLARED_ALLOCATION,
                        "aggregate_estimated_value": decimal_text(estimated),
                    }
                )
            checked_rows.append(checked)
    return findings, warnings, checked_rows, not_checked_rows


def _deviation_values(
    estimated: Decimal | None,
    control: Decimal | None,
) -> tuple[Decimal | None, Decimal | None]:
    if estimated is None or control is None:
        return None, None
    absolute = abs(estimated - control)
    if control == 0:
        percent = Decimal(0) if absolute == 0 else None
    else:
        percent = absolute / abs(control) * Decimal(100)
    return absolute, percent


def _deviation_exceeds(
    estimated: Decimal | None,
    control: Decimal | None,
    tolerance: Decimal,
) -> bool:
    absolute, percent = _deviation_values(estimated, control)
    if absolute is None:
        return False
    if control == 0:
        return absolute != 0
    return percent is not None and percent > tolerance


def _price_finding(
    kind: str,
    row: dict[str, Any],
    catalog_item: dict[str, Any],
    *,
    estimated: Decimal,
    control: Decimal,
    quantity: Decimal | None,
    estimate_total: Decimal | None,
    mcp_total: Decimal | None,
) -> dict[str, Any]:
    deviation_absolute, deviation_percent = _deviation_values(estimated, control)
    if kind in {"price_overstatement", "price_understatement"}:
        impact = (estimated - control) * quantity if quantity is not None else None
        impact_basis = "unit_price_difference_at_estimate_quantity"
        impact_formula = "(estimate_unit_price - mcp_unit_price) * estimate_quantity"
    else:
        impact = estimated - control
        impact_basis = "estimate_total_vs_estimate_quantity_at_mcp_price"
        impact_formula = "estimate_total - (estimate_quantity * mcp_unit_price)"
    result = {
        "type": kind,
        "source_row": row["source_row"],
        "source_position": row.get("position"),
        "source_rows": [row["source_row"]],
        "original_room_name": row["room"],
        "canonical_room_id": None,
        "canonical_room_name": "—",
        "original_work_name": row["work_name"],
        "canonical_work_name": catalog_item["name"],
        "mcp_work_id": catalog_item["id"],
        "mcp_work_name": catalog_item["name"],
        "estimated_value": decimal_text(estimated),
        "control_value": decimal_text(control),
        "unit": catalog_item["unit"],
        "estimate_price": row.get("price"),
        "mcp_price": catalog_item["price"],
        "quantity": decimal_text(quantity),
        "estimate_total": decimal_text(estimate_total),
        "mcp_total": decimal_text(mcp_total),
        "deviation_absolute": decimal_text(deviation_absolute),
        "deviation_percent": decimal_text(deviation_percent),
        "financial_impact": {
            "status": "calculated" if impact is not None else "not_available",
            "basis": impact_basis,
            "formula": impact_formula,
            "signed_value": decimal_text(impact),
            "absolute_value": decimal_text(abs(impact)) if impact is not None else None,
        },
        "calculation_trace_refs": [f"price:row:{row['source_row']}:{catalog_item['id']}"],
        "severity": "high" if deviation_percent is None or deviation_percent > Decimal(20) else "warning",
        "status": "preliminary",
    }
    result["safe_summary"] = (
        f"По строке {row['source_row']} выявлено предварительное расхождение стоимости с "
        f"контрольным расчётом по позиции MCP {catalog_item['id']}; строка требует проверки специалистом."
    )
    return result


def run_price_checks(
    estimate: dict[str, Any],
    mapping: dict[str, Any],
    price_catalog: dict[str, Any],
    tolerance: Decimal,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """Re-resolve every mapped id in the saved catalog and calculate prices in Python."""
    catalog_by_id = {item["id"]: item for item in price_catalog["items"]}
    price_map = {item["source_row"]: item for item in mapping["price_matches"]}
    unsupported = {item["source_row"]: item for item in mapping["price_unsupported"]}
    checks: list[dict[str, Any]] = []
    findings: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    trace_entries: list[dict[str, Any]] = []

    for row in estimate["rows"]:
        source_row = row["source_row"]
        if source_row in unsupported:
            item = unsupported[source_row]
            warnings.append(
                {
                    "code": "unsupported_price_work",
                    "source_row": source_row,
                    "work_name": row["work_name"],
                    "message": item["reason"],
                }
            )
            checks.append(
                {
                    "source_row": source_row,
                    "estimate_work": row["work_name"],
                    "mcp_work_id": None,
                    "mcp_work_name": None,
                    "unit": row["unit"],
                    "estimate_price": row["price"],
                    "mcp_price": None,
                    "quantity": row["quantity"],
                    "estimate_total": row["total"],
                    "total_source": row.get("total_source"),
                    "mcp_total": None,
                    "price_deviation_absolute": None,
                    "price_deviation_percent": None,
                    "total_deviation_absolute": None,
                    "total_deviation_percent": None,
                    "deviation_absolute": None,
                    "deviation_percent": None,
                    "status": "not_checked",
                    "reason": "unsupported_price_work",
                }
            )
            continue

        price_mapping = price_map.get(source_row)
        if price_mapping is None:
            warnings.append(
                {
                    "code": "missing_price_mapping",
                    "source_row": source_row,
                    "message": "Ценовой mapping отсутствует.",
                }
            )
            checks.append(
                {
                    "source_row": source_row,
                    "source_position": row.get("position"),
                    "estimate_work": row["work_name"],
                    "mcp_work_id": None,
                    "mcp_work_name": None,
                    "unit": row["unit"],
                    "estimate_price": row["price"],
                    "mcp_price": None,
                    "quantity": row["quantity"],
                    "estimate_total": row["total"],
                    "total_source": row.get("total_source"),
                    "mcp_total": None,
                    "price_deviation_absolute": None,
                    "price_deviation_percent": None,
                    "total_deviation_absolute": None,
                    "total_deviation_percent": None,
                    "deviation_absolute": None,
                    "deviation_percent": None,
                    "status": "not_checked",
                    "reason": "missing_price_mapping",
                }
            )
            continue

        mcp_work_id = price_mapping["mcp_work_id"]
        catalog_item = catalog_by_id.get(mcp_work_id)
        if catalog_item is None:
            raise AuditError(
                "price_catalog_lookup_failed",
                "Позиция MCP из mapping отсутствует в сохранённом каталоге.",
                {"source_row": source_row, "mcp_work_id": mcp_work_id},
            )
        quantity = _decimal(row.get("quantity"))
        estimate_price = _decimal(row.get("price"))
        estimate_total = _decimal(row.get("total"))
        total_source = str(row.get("total_source") or "provided_or_cached")
        source_total_available = total_source == "provided_or_cached"
        mcp_price = Decimal(catalog_item["price"])
        mcp_total = quantity * mcp_price if quantity is not None else None
        price_absolute, price_percent = _deviation_values(estimate_price, mcp_price)
        total_absolute, total_percent = _deviation_values(estimate_total, mcp_total)
        unit_price_impact = (
            (estimate_price - mcp_price) * quantity
            if estimate_price is not None and quantity is not None
            else None
        )
        total_cost_impact = (
            estimate_total - mcp_total
            if estimate_total is not None and mcp_total is not None
            else None
        )
        trace_id = f"price:row:{row['source_row']}:{catalog_item['id']}"
        trace_entries.append(
            {
                "trace_id": trace_id,
                "source_row": source_row,
                "metric": "price_and_total_check",
                "formula": {
                    "mcp_total": "estimate_quantity * mcp_unit_price",
                    "unit_price_impact": "(estimate_unit_price - mcp_unit_price) * estimate_quantity",
                    "total_cost_impact": "estimate_total - mcp_total",
                },
                "inputs": {
                    "estimate_quantity": decimal_text(quantity),
                    "estimate_unit_price": decimal_text(estimate_price),
                    "mcp_unit_price": catalog_item["price"],
                    "estimate_total": decimal_text(estimate_total),
                    "estimate_total_source": total_source,
                },
                "results": {
                    "mcp_total": decimal_text(mcp_total),
                    "unit_price_impact": decimal_text(unit_price_impact),
                    "total_cost_impact": decimal_text(total_cost_impact),
                },
                "warnings": [],
            }
        )
        row_findings: list[dict[str, Any]] = []
        if estimate_price is not None and _deviation_exceeds(estimate_price, mcp_price, tolerance):
            row_findings.append(
                _price_finding(
                    "price_overstatement" if estimate_price > mcp_price else "price_understatement",
                    row,
                    catalog_item,
                    estimated=estimate_price,
                    control=mcp_price,
                    quantity=quantity,
                    estimate_total=estimate_total,
                    mcp_total=mcp_total,
                )
            )
        # Отдельный finding по итогу нужен только без цены за единицу. Когда цена есть,
        # ошибку итога уже показывает arithmetic_mismatch, а сравнение с MCP ниже
        # остаётся финансовым результатом и не дублирует замечание.
        if (
            source_total_available
            and estimate_price is None
            and estimate_total is not None
            and mcp_total is not None
            and _deviation_exceeds(estimate_total, mcp_total, tolerance)
        ):
            row_findings.append(
                _price_finding(
                    "total_cost_overstatement" if estimate_total > mcp_total else "total_cost_understatement",
                    row,
                    catalog_item,
                    estimated=estimate_total,
                    control=mcp_total,
                    quantity=quantity,
                    estimate_total=estimate_total,
                    mcp_total=mcp_total,
                )
            )
        findings.extend(row_findings)

        missing: list[str] = []
        if quantity is None:
            missing.append("quantity")
            warnings.append(
                {"code": "price_check_missing_quantity", "source_row": source_row, "message": "Для пересчёта стоимости отсутствует количество."}
            )
        if estimate_price is None:
            missing.append("estimate_price")
            warnings.append(
                {"code": "price_check_missing_price", "source_row": source_row, "message": "В смете отсутствует цена за единицу."}
            )
        if estimate_total is None:
            missing.append("estimate_total")
            warnings.append(
                {"code": "price_check_missing_total", "source_row": source_row, "message": "В смете отсутствует стоимость строки."}
            )
        elif not source_total_available:
            missing.append("source_total")
            warnings.append({
                "code": "price_check_imputed_total",
                "level": "info",
                "source_row": source_row,
                "message": "Стоимость строки отсутствовала в XLSX и рассчитана как количество × цена сметы.",
            })
        comparisons = int(estimate_price is not None) + int(
            source_total_available and estimate_total is not None and mcp_total is not None
        )
        has_arithmetic_mismatch = any(
            issue.get("type") == "arithmetic_mismatch" for issue in row.get("issues", [])
        )
        if row_findings or has_arithmetic_mismatch:
            status = "deviation_found"
        elif comparisons == 2:
            status = "checked"
        elif comparisons == 1:
            status = "partially_checked"
        else:
            status = "not_checked"
        primary_absolute = total_absolute if source_total_available and total_absolute is not None else price_absolute
        primary_percent = total_percent if source_total_available and total_percent is not None else price_percent
        checks.append(
            {
                "source_row": source_row,
                "source_position": row.get("position"),
                "estimate_work": row["work_name"],
                "mcp_work_id": catalog_item["id"],
                "mcp_work_name": catalog_item["name"],
                "unit": catalog_item["unit"],
                "estimate_price": decimal_text(estimate_price),
                "mcp_price": catalog_item["price"],
                "quantity": decimal_text(quantity),
                "estimate_total": decimal_text(estimate_total),
                "total_source": total_source,
                "mcp_total": decimal_text(mcp_total),
                "price_deviation_absolute": decimal_text(price_absolute),
                "price_deviation_percent": decimal_text(price_percent),
                "total_deviation_absolute": decimal_text(total_absolute),
                "total_deviation_percent": decimal_text(total_percent),
                "unit_price_impact": decimal_text(unit_price_impact),
                "unit_price_impact_formula": "(estimate_price - mcp_price) * estimate_quantity",
                "total_cost_impact": decimal_text(total_cost_impact),
                "total_cost_impact_formula": "estimate_total - (estimate_quantity * mcp_price)",
                "deviation_absolute": decimal_text(primary_absolute),
                "deviation_percent": decimal_text(primary_percent),
                "status": status,
                "reason": ",".join(missing) if missing else None,
            }
        )
    return checks, findings, warnings, trace_entries


def enrich_findings(
    findings: list[dict[str, Any]],
    price_checks: list[dict[str, Any]],
    quantity_checks: list[dict[str, Any]],
) -> None:
    """Add stable report IDs and conservative, unsigned-currency impact metadata."""
    checks_by_row = {item.get("source_row"): item for item in price_checks}
    quantity_by_row = {item.get("source_row"): item for item in quantity_checks}
    quantity_types = {"quantity_overstatement", "quantity_understatement"}
    for index, finding in enumerate(findings, start=1):
        finding["finding_id"] = f"finding_{index:03d}"
        if finding.get("type") in {"price_overstatement", "price_understatement"}:
            source_row = finding.get("source_row")
            price_check = checks_by_row.get(source_row, {})
            quantity_check = quantity_by_row.get(source_row, {})
            estimate_quantity = _decimal(price_check.get("quantity"))
            control_quantity = _decimal(quantity_check.get("control_value"))
            estimate_price = _decimal(price_check.get("estimate_price"))
            mcp_price = _decimal(price_check.get("mcp_price"))
            estimate_total = _decimal(price_check.get("estimate_total"))
            if all(value is not None for value in (
                estimate_quantity, control_quantity, estimate_price, mcp_price, estimate_total,
            )) and len(quantity_check.get("source_rows") or [source_row]) == 1:
                reference_total = control_quantity * mcp_price
                quantity_effect = (estimate_quantity - control_quantity) * mcp_price
                price_effect = (estimate_price - mcp_price) * estimate_quantity
                arithmetic_effect = estimate_total - estimate_quantity * estimate_price
                full_variance = estimate_total - reference_total
                finding["line_cost_analysis"] = {
                    "status": "calculated",
                    "estimate_total": decimal_text(estimate_total),
                    "reference_total": decimal_text(reference_total),
                    "full_variance_signed": decimal_text(full_variance),
                    "full_variance_absolute": decimal_text(abs(full_variance)),
                    "full_variance_percent": decimal_text(
                        abs(full_variance) / abs(reference_total) * Decimal(100)
                    ) if reference_total != 0 else None,
                    "quantity_effect_signed": decimal_text(quantity_effect),
                    "price_effect_signed": decimal_text(price_effect),
                    "arithmetic_effect_signed": decimal_text(arithmetic_effect),
                    "decomposition_formula": (
                        "estimate_total - control_quantity * mcp_price = "
                        "(estimate_quantity - control_quantity) * mcp_price + "
                        "(estimate_price - mcp_price) * estimate_quantity + "
                        "(estimate_total - estimate_quantity * estimate_price)"
                    ),
                    "simultaneous_quantity_and_price_deviation": bool(
                        quantity_check.get("threshold_exceeded")
                    ),
                    "total_source": price_check.get("total_source"),
                }
        if finding.get("type") not in quantity_types:
            finding.setdefault(
                "financial_impact",
                {
                    "status": "not_applicable",
                    "basis": None,
                    "formula": None,
                    "signed_value": None,
                    "absolute_value": None,
                },
            )
            continue
        source_rows = finding.get("source_rows") or [finding.get("source_row")]
        row_prices = [
            checks_by_row.get(source_row, {}).get("mcp_price") for source_row in source_rows
        ]
        reference_prices = {price for price in row_prices if price is not None}
        estimated = _decimal(finding.get("estimated_value"))
        control = _decimal(finding.get("control_value"))
        if (
            len(row_prices) == len(source_rows)
            and all(price is not None for price in row_prices)
            and len(reference_prices) == 1
            and estimated is not None
            and control is not None
        ):
            reference_price = Decimal(next(iter(reference_prices)))
            impact = (estimated - control) * reference_price
            finding["financial_impact"] = {
                "status": "calculated",
                "basis": "quantity_difference_at_mcp_unit_price",
                "formula": "(estimate_quantity - control_quantity) * mcp_unit_price",
                "reference_unit_price": decimal_text(reference_price),
                "signed_value": decimal_text(impact),
                "absolute_value": decimal_text(abs(impact)),
            }
        else:
            finding["financial_impact"] = {
                "status": "not_available",
                "basis": "quantity_difference_at_mcp_unit_price",
                "formula": "(estimate_quantity - control_quantity) * mcp_unit_price",
                "reference_unit_price": None,
                "signed_value": None,
                "absolute_value": None,
                "reason": "single_mcp_reference_price_unavailable",
            }


def audit_summary(
    estimate: dict[str, Any],
    findings: list[dict[str, Any]],
    warnings: list[dict[str, Any]],
    checked_rows: list[dict[str, Any]],
    not_checked_rows: list[dict[str, Any]],
    price_checks: list[dict[str, Any]],
    quantity_checks: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    quantity_checks = quantity_checks or []
    total = len(estimate.get("rows", []))
    quantity_checked = len({item.get("source_row") for item in checked_rows})
    quantity_not_checked = len({item.get("source_row") for item in not_checked_rows})
    price_full = 0
    price_partial = 0
    price_not_checked = 0
    price_deviations = 0
    quantity_exact = len({
        item.get("source_row") for item in quantity_checks if item.get("status") == "exact_match"
    })
    quantity_below = len({
        item.get("source_row") for item in quantity_checks if item.get("status") == "below_threshold"
    })
    quantity_deviations = len({
        item.get("source_row") for item in quantity_checks if item.get("status") == "deviation_found"
    })
    for check in price_checks:
        comparisons = int(check.get("estimate_price") is not None and check.get("mcp_price") is not None)
        comparisons += int(
            check.get("total_source") == "provided_or_cached"
            and check.get("estimate_total") is not None
            and check.get("mcp_total") is not None
        )
        if comparisons == 2:
            price_full += 1
        elif comparisons == 1:
            price_partial += 1
        else:
            price_not_checked += 1
        if check.get("status") == "deviation_found":
            price_deviations += 1
    by_severity: dict[str, int] = {"high": 0, "warning": 0}
    for finding in findings:
        severity = str(finding.get("severity") or "warning")
        by_severity[severity] = by_severity.get(severity, 0) + 1
    price_covered = price_full + price_partial
    door_installation_rows = [
        item
        for item in checked_rows
        if item.get("quantity_check_scope") == DOOR_QUANTITY_SCOPE
    ]
    door_check = door_installation_rows[0] if door_installation_rows else {}
    return {
        "estimate_rows": total,
        "checked_rows": quantity_checked,
        "not_checked_rows": quantity_not_checked,
        "quantity_checked_rows": quantity_checked,
        "quantity_not_checked_rows": quantity_not_checked,
        "quantity_coverage_percent": decimal_text(
            Decimal(quantity_checked) / Decimal(total) * Decimal(100) if total else Decimal(0)
        ),
        "quantity_exact_match_rows": quantity_exact,
        "quantity_below_threshold_rows": quantity_below,
        "quantity_deviation_rows": quantity_deviations,
        "price_fully_checked_rows": price_full,
        "price_partially_checked_rows": price_partial,
        "price_not_checked_rows": price_not_checked,
        "price_coverage_percent": decimal_text(
            Decimal(price_covered) / Decimal(total) * Decimal(100) if total else Decimal(0)
        ),
        "price_deviation_rows": price_deviations,
        "door_installation_rows": len(door_installation_rows),
        "door_installation_check_scope": DOOR_QUANTITY_SCOPE if door_installation_rows else None,
        "door_installation_allocation_status": (
            ESTIMATE_DECLARED_ALLOCATION if door_installation_rows else None
        ),
        "door_installation_estimate_total": door_check.get("aggregate_estimated_value"),
        "door_installation_unique_plan_total": door_check.get("control_value"),
        "findings_count": len(findings),
        "findings_by_severity": by_severity,
        "warnings_count": sum(item.get("level") != "info" for item in warnings),
        "info_count": sum(item.get("level") == "info" for item in warnings),
        "completion_status": (
            "completed_partial"
            if quantity_not_checked or price_partial or price_not_checked
            else "completed"
        ),
    }


def _warning_markdown(warning: dict[str, Any]) -> str:
    message = str(warning.get("message") or "Предупреждение без текстового описания.")
    details = []
    if warning.get("source_row") is not None:
        details.append(f"строка {warning['source_row']}")
    if warning.get("room_id"):
        details.append(f"помещение {warning['room_id']}")
    return f"{message} ({', '.join(details)})" if details else message


def _ru_number(value: Any) -> str:
    if value is None or value == "":
        return "—"
    return str(value).replace(".", ",")


def _ru_percent(value: Any) -> str:
    if value is None or value == "":
        return "—"
    rounded = Decimal(str(value)).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
    text = format(rounded, "f").rstrip("0").rstrip(".")
    return f"{text.replace('.', ',')}%"


def _finding_type_label(value: Any) -> str:
    return {
        "quantity_overstatement": "выше контроля",
        "quantity_understatement": "ниже контроля",
        "exact_duplicate": "дубликат",
        "arithmetic_mismatch": "ошибка стоимости",
        "invalid_quantity": "некорректное количество",
        "price_overstatement": "цена выше MCP",
        "price_understatement": "цена ниже MCP",
        "total_cost_overstatement": "стоимость выше MCP",
        "total_cost_understatement": "стоимость ниже MCP",
    }.get(str(value), str(value or "расхождение"))


def build_audit_summary_markdown(
    summary: dict[str, Any],
    findings: list[dict[str, Any]],
    warnings: list[dict[str, Any]],
    geometry: dict[str, Any],
    report_path: Path,
    visual_insights: dict[str, Any] | None = None,
    llm_insights: dict[str, Any] | None = None,
) -> str:
    lines = [
        "# Результаты предварительной проверки сметы",
        "",
        (
            "**Статус:** завершён частично — часть строк не проверена из-за отсутствующих данных"
            if summary.get("completion_status") == "completed_partial"
            else "**Статус:** завершён"
        ),
        "",
        "| Строк всего | Проверено | Не проверено | Высокая важность | Остальные расхождения |",
        "|---:|---:|---:|---:|---:|",
        (
            f"| {summary['estimate_rows']} | {summary['checked_rows']} | "
            f"{summary['not_checked_rows']} | "
            f"{summary.get('findings_by_severity', {}).get('high', 0)} | "
            f"{summary.get('findings_by_severity', {}).get('warning', 0)} |"
        ),
        "",
        "## Итог",
        "",
    ]
    if summary["findings_count"] == 0:
        lines.append("По проверенным строкам предварительных расхождений не выявлено")
    else:
        lines.append(
            f"Выявлено {summary['findings_count']} предварительных расхождений; связанные "
            "строки требуют проверки специалистом."
        )
    if summary["not_checked_rows"]:
        lines.append(f"{summary['not_checked_rows']} строк требуют проверки специалистом.")
    if summary.get("door_installation_rows"):
        lines.append(
            "Установка дверей проверена только суммарно по объекту: распределение между "
            "помещениями принято из сметы и отдельно по комнатам не проверялось. "
            f"Сумма по смете — {_ru_number(summary.get('door_installation_estimate_total'))}; "
            "уникальных дверей на плане — "
            f"{_ru_number(summary.get('door_installation_unique_plan_total'))}."
        )
    missing_window_heights = any(
        window["height_m"] is None
        for room in geometry["rooms"]
        for window in room["windows"]
    )
    missing_wall_area = any(
        warning.get("code") == "missing_geometry"
        and warning.get("metric") == "net_wall_area_m2"
        for warning in warnings
    )
    if missing_window_heights and missing_wall_area:
        lines.append(
            "Причина непроверенных стеновых работ: отсутствуют высоты окон, поэтому нельзя "
            "рассчитать площадь окон, проёмов и чистую площадь стен."
        )
    lines.extend(["", "## Предварительные расхождения", ""])
    if findings:
        lines.extend(
            [
                "| Строка | Помещение | Работа | Смета | Контроль | Отклонение | Статус |",
                "|---:|---|---|---:|---:|---:|---|",
            ]
        )
        for finding in findings:
            unit = (
                ""
                if finding.get("type") in report.PRICE_TYPES | report.TOTAL_TYPES
                else str(finding.get("unit") or "")
            )
            estimated = f"{_ru_number(finding.get('estimated_value'))} {unit}".strip()
            control = f"{_ru_number(finding.get('control_value'))} {unit}".strip()
            deviation = _ru_percent(finding.get("deviation_percent"))
            lines.append(
                "| " + " | ".join(
                    _markdown_cell(value)
                    for value in (
                        finding.get("source_row"),
                        (
                            "Весь объект"
                            if finding.get("quantity_check_scope") == DOOR_QUANTITY_SCOPE
                            else finding.get("original_room_name")
                        ),
                        finding.get("original_work_name"),
                        estimated,
                        control,
                        deviation,
                        _finding_type_label(finding.get("type")),
                    )
                ) + " |"
            )
    else:
        lines.append("- нет")

    lines.extend(["", "## Технические предупреждения и непроверенные данные", ""])
    user_warnings = [warning for warning in warnings if warning.get("level") != "info"]
    if user_warnings:
        for warning in user_warnings[:8]:
            lines.append(f"- {_warning_markdown(warning)}")
        if len(user_warnings) > 8:
            lines.append(f"- Ещё предупреждений: {len(user_warnings) - 8}. Они доступны в HTML-отчёте.")
    else:
        lines.append("- нет")
    if visual_insights and visual_insights.get("status") != "skipped":
        items = visual_insights.get("items", [])
        lines.extend(
            [
                "",
                "## Наблюдения по фотографиям",
                "",
                (
                    f"Проанализировано фотографий: {visual_insights.get('photos_count', 0)}. "
                    "Фотографии не привязаны к помещениям; строки сметы показывают только "
                    "позиции соответствующей работы."
                ),
            ]
        )
        by_work: dict[str, dict[str, Any]] = {}
        quality_count = 0
        for item in items:
            if item.get("category") == "quality":
                quality_count += 1
                continue
            work = str(item.get("estimate_work") or "").strip()
            if not work:
                continue
            target = by_work.setdefault(
                work,
                {"source_rows": item.get("source_rows", []), "statuses": {}},
            )
            status = str(item.get("status") or "")
            target["statuses"].setdefault(status, set()).add(item.get("photo_id"))
        status_labels = {
            "observed": "видно",
            "not_observed": "не видно",
            "not_assessable": "не поддаётся оценке",
        }
        for work in SUPPORTED_WORKS:
            target = by_work.get(work)
            if target is None:
                continue
            rows = ", ".join(str(row) for row in target["source_rows"])
            counts = "; ".join(
                f"{label} на {len(target['statuses'][status])} фото"
                for status, label in status_labels.items()
                if target["statuses"].get(status)
            )
            lines.append(f"- {work} (строки {rows}): {counts}.")
        lines.append(f"- Признаков качества для очной проверки: {quality_count}.")
    if llm_insights and llm_insights.get("status") == "generated":
        lines.extend(["", "## Аналитические гипотезы", "", str(llm_insights.get("summary") or "")])
        for item in llm_insights.get("items", [])[:3]:
            lines.append(f"- {item.get('title')}: {item.get('observation')}")
    lines.extend(
        [
            "",
            "## HTML-отчёт",
            "",
            f"`{report_path}`",
            "",
            f"Дисклеймер: {report.DISCLAIMER}.",
            "",
            "## Дополнительно",
            "",
            (
                "Могу сформировать отдельную предварительную XLSX-смету по подтверждённой "
                "геометрии и сохранённым ценам MCP. Для этого попросите сформировать смету "
                "отдельным новым сообщением."
            ),
        ]
    )
    return "\n".join(lines)


def render_audit_summary(api: Any, job_id: str) -> dict[str, Any]:
    root = Path(api.skill_job_dir(validate_job_id(job_id))).resolve(strict=False)
    manifest = load_manifest(root)
    if manifest.get("audit_status") == "llm_insights_required":
        context_package = read_json(
            output_path(root, "llm_context.json"), code="llm_insights_context_required"
        )
        context_sha256 = canonical_json_sha256(context_package)
        if context_sha256 != manifest.get("llm_context_sha256"):
            raise AuditError(
                "llm_context_changed",
                "Сохранённый контекст аналитического этапа изменился.",
            )
        visual_artifact = read_json(
            output_path(root, "visual_insights.json"), code="visual_review_required"
        )
        if canonical_json_sha256(visual_artifact) != manifest.get("visual_insights_sha256"):
            raise AuditError(
                "visual_insights_changed",
                "Visual insights изменились после подготовки LLM-контекста.",
            )
        findings_artifact = read_json(
            output_path(root, "findings.json"), code="audit_result_required"
        )
        result = {
            "status": "llm_insights_required",
            "llm_insights_delegation": insights.delegation(context_package),
            "summary": findings_artifact.get("summary", {}),
            "visual_summary": {
                "status": visual_artifact.get("status"),
                "photos_count": visual_artifact.get("photos_count", 0),
                "insights_count": len(visual_artifact.get("items", [])),
            },
            "llm_context_sha256": context_sha256,
            "next_action": "schedule_llm_insights_subagent",
            "assistant_instruction": LLM_INSIGHTS_ASSISTANT_INSTRUCTION,
        }
        if len(canonical_json_text(result)) > 14_500:
            raise AuditError(
                "llm_context_too_large",
                "Сохранённый пакет аналитического этапа превышает безопасный transport-размер.",
            )
        return result
    if not manifest.get("audit_completed") or manifest.get("audit_status") != "completed":
        raise AuditError("audit_not_completed", "Итоговый аудит ещё не завершён.")

    estimate = read_json(output_path(root, "estimate_normalized.json"), code="estimate_required")
    geometry = read_json(output_path(root, "geometry.json"), code="geometry_required")
    artifact = read_json(output_path(root, "findings.json"), code="audit_result_required")
    if not isinstance(estimate, dict) or not isinstance(geometry, dict) or not isinstance(artifact, dict):
        raise AuditError("storage_error", "Сохранённый результат аудита повреждён.")
    findings = artifact.get("findings")
    warnings = artifact.get("warnings")
    checked_rows = artifact.get("checked_rows")
    not_checked_rows = artifact.get("not_checked_rows")
    rows = estimate.get("rows")
    if not all(isinstance(value, list) for value in (rows, findings, warnings, checked_rows, not_checked_rows)):
        raise AuditError("storage_error", "Сохранённый результат аудита повреждён.")

    report_path = output_path(root, "report.html")
    if not report_path.is_file():
        raise AuditError("report_required", "HTML-отчёт аудита не найден.")
    price_checks = artifact.get("price_checks")
    if not isinstance(price_checks, list):
        price_checks = []
    quantity_checks = artifact.get("quantity_checks")
    if not isinstance(quantity_checks, list):
        quantity_checks = []
    summary = artifact.get("summary")
    if not isinstance(summary, dict):
        summary = audit_summary(
            estimate, findings, warnings, checked_rows, not_checked_rows, price_checks,
            quantity_checks,
        )
    visual_path = output_path(root, "visual_insights.json")
    insights_path = output_path(root, "llm_insights.json")
    visual_insights = (
        read_json(visual_path, code="visual_review_required")
        if visual_path.is_file()
        else visual.empty_artifact()
    )
    llm_insights = (
        read_json(insights_path, code="llm_insights_required")
        if insights_path.is_file()
        else {"status": "no_useful_observations", "summary": "", "items": []}
    )
    return {
        "status": "audit_completed",
        "summary": summary,
        "report_artifact": {
            "path": str(report_path),
            "size_bytes": report_path.stat().st_size,
            "sha256": sha256_file(report_path),
        },
        "audit_summary_markdown": build_audit_summary_markdown(
            summary, findings, warnings, geometry, report_path, visual_insights, llm_insights
        ),
        "next_action": "display_audit_summary_markdown_verbatim",
        "assistant_instruction": AUDIT_SUMMARY_ASSISTANT_INSTRUCTION,
    }


def run_audit(
    api: Any,
    ctx: Any,
    job_id: str,
    mapping_task_id: Any,
    mapping_payload: Any,
    tolerance_percent: Any = 5,
) -> tuple[dict[str, Any], dict[str, Any]]:
    root = job_dir(api, job_id)
    manifest = load_manifest(root)
    if not isinstance(mapping_task_id, str) or not mapping_task_id.strip() or len(mapping_task_id.strip()) > 200:
        raise AuditError("mapping_execution_failed", "mapping_task_id должен быть непустой строкой до 200 символов.")
    normalized_task_id = mapping_task_id.strip()
    if not manifest.get("geometry_confirmed"):
        raise AuditError("geometry_confirmation_required", "Сначала подтвердите geometry отдельным новым сообщением пользователя.")
    geometry = read_json(output_path(root, "geometry.json"), code="geometry_required")
    _require_consistent_shared_doors(geometry)
    if geometry_sha256(geometry) != manifest.get("confirmed_geometry_sha256"):
        raise AuditError("geometry_changed_since_confirmation", "Подтверждённый hash geometry не совпадает с текущим.")
    if manifest.get("geometry_confirmed_revision") != manifest.get("geometry_revision"):
        raise AuditError(
            "stale_geometry_confirmation",
            "Подтверждение относится не к текущей geometry revision.",
        )
    derived_issues = vision.derived_state_issues(geometry)
    if derived_issues:
        raise AuditError(
            "geometry_derived_state_incomplete",
            "Для части помещений есть исходные размеры, но не рассчитаны производные значения; аудит не запущен.",
            {"incomplete_fields": derived_issues},
        )
    catalog_path = output_path(root, "price_catalog.json")
    if not catalog_path.is_file():
        raise AuditError(
            "price_catalog_required",
            f"После подтверждения geometry вызовите {PRICE_CATALOG_TOOL} и сохраните его result.",
        )
    price_catalog = read_json(catalog_path, code="price_catalog_required")
    # Mapping возвращает только ID. Цену перечитываем из сохранённого каталога
    # и сверяем его hash, поэтому модель не может подменить число в своём ответе.
    if not isinstance(price_catalog, dict) or canonical_json_sha256(price_catalog) != manifest.get("price_catalog_sha256"):
        raise AuditError("price_catalog_changed", "Сохранённый каталог MCP изменился после validation.")
    expected_token = manifest.get("mapping_generation_token")
    supplied_token = mapping_payload.get("delegation_token") if isinstance(mapping_payload, dict) else None
    if not isinstance(expected_token, str) or not expected_token:
        raise AuditError(
            "mapping_delegation_invalid",
            "Для текущей подтверждённой geometry revision отсутствует Mapping-делегация.",
            {"reason": "mapping_delegation_not_issued"},
        )
    if not isinstance(supplied_token, str) or not secrets.compare_digest(supplied_token, expected_token):
        raise AuditError(
            "mapping_delegation_invalid",
            "Mapping получен не из делегации для текущей подтверждённой geometry revision.",
            {
                "reason": "missing_delegation_token"
                if not isinstance(supplied_token, str)
                else "stale_or_foreign_delegation_token"
            },
        )
    estimate = read_json(output_path(root, "estimate_normalized.json"), code="estimate_required")
    tolerance = _decimal(tolerance_percent)
    if tolerance is None or tolerance < 0 or tolerance > 100:
        raise AuditError("invalid_tolerance", "tolerance_percent должен быть числом от 0 до 100.")
    try:
        mapping = validate_mapping(mapping_payload, estimate, geometry, expected_token, price_catalog)
    except AuditError as exc:
        if exc.code not in {"invalid_mapping", "incompatible_unit"}:
            raise
        validation_error = {"message": exc.message}
        if exc.details:
            validation_error["details"] = exc.details
        raise AuditError(
            "mapping_schema_invalid",
            "Результат Mapping-субагента не прошёл validation.",
            {
                "validation_errors": [validation_error],
                "allowed_next_action": "rerun_mapping_subagent",
            },
        ) from exc
    mapping_fingerprint = canonical_json_sha256(mapping)
    manifest["mapping_task_id"] = normalized_task_id
    manifest["mapping_sha256"] = mapping_fingerprint
    manifest["mapping_schema_version"] = mapping["schema_version"]
    manifest["mapping_validation_status"] = "validated"
    unresolved = [
        *mapping["room_unresolved"],
        *mapping["work_unresolved"],
        *mapping["price_unresolved"],
    ]
    # При неоднозначном соответствии считать ещё рано: частичный Mapping создал бы
    # внешне полный отчёт с произвольно пропущенными строками.
    if unresolved:
        manifest["audit_status"] = "mapping_review_required"
        save_manifest(root, manifest)
        return manifest, {"status": "mapping_review_required", "unresolved": unresolved}

    quantities, trace, calculation_warnings = calculate_quantities(geometry)
    quantity_checks: list[dict[str, Any]] = []
    findings, warnings, checked_rows, not_checked_rows = run_checks(
        estimate, geometry, mapping, quantities, trace, tolerance, quantity_checks
    )
    price_checks, price_findings, price_warnings, price_trace_entries = run_price_checks(
        estimate, mapping, price_catalog, tolerance
    )
    trace["entries"].extend(price_trace_entries)
    findings.extend(price_findings)
    warnings.extend(price_warnings)
    warnings.extend(calculation_warnings)
    warning_seen: set[str] = set()
    warnings = [
        item for item in warnings
        if not (json.dumps(item, ensure_ascii=False, sort_keys=True) in warning_seen)
        and not warning_seen.add(json.dumps(item, ensure_ascii=False, sort_keys=True))
    ]
    enrich_findings(findings, price_checks, quantity_checks)
    summary = audit_summary(
        estimate, findings, warnings, checked_rows, not_checked_rows, price_checks,
        quantity_checks,
    )
    findings_artifact = {
        "schema_version": 1,
        "tolerance_percent": decimal_text(tolerance),
        "price_catalog_sha256": manifest["price_catalog_sha256"],
        "findings": findings,
        "warnings": warnings,
        "checked_rows": checked_rows,
        "not_checked_rows": not_checked_rows,
        "quantity_checks": quantity_checks,
        "price_checks": price_checks,
        "summary": summary,
    }
    atomic_write_json(output_path(root, "mapping.json"), mapping)
    atomic_write_json(output_path(root, "quantities.json"), quantities)
    atomic_write_json(output_path(root, "calculation_trace.json"), trace)
    atomic_write_json(
        output_path(root, "price_checks.json"),
        {
            "schema_version": 1,
            "tolerance_percent": decimal_text(tolerance),
            "price_catalog_sha256": manifest["price_catalog_sha256"],
            "checks": price_checks,
        },
    )
    atomic_write_json(output_path(root, "findings.json"), findings_artifact)
    # Детерминированные расчёты уже готовы, но аудит ещё не завершён: сначала
    # пользователь выбирает фото, затем обязательный аналитический субагент.
    manifest["audit_completed"] = False
    manifest["report_generated"] = False
    manifest["audit_status"] = "visual_review_required"
    manifest["deterministic_audit_prepared_at"] = now_iso()
    for key in (
        "llm_context_sha256", "llm_insights_task_id", "llm_insights_sha256",
        "llm_insights_status", "llm_insights_completed_at", "llm_insights_attempts",
    ):
        manifest.pop(key, None)
    save_manifest(root, manifest)
    return manifest, {
        "status": "visual_review_required",
        "summary": summary,
        "visual_review_offer_markdown": VISUAL_REVIEW_OFFER_MARKDOWN,
        "next_action": "await_site_photo_zip_or_skip",
        "assistant_instruction": VISUAL_REVIEW_OFFER_ASSISTANT_INSTRUCTION,
    }


def _visual_estimate_works(
    estimate: dict[str, Any],
    mapping: dict[str, Any],
) -> list[dict[str, Any]]:
    rows_by_id = {row["source_row"]: row for row in estimate.get("rows", [])}
    grouped: dict[str, dict[str, Any]] = {}
    for item in mapping.get("work_matches", []):
        canonical = item.get("canonical_work")
        source_row = item.get("source_row")
        row = rows_by_id.get(source_row)
        if canonical not in SUPPORTED_WORKS or row is None:
            continue
        target = grouped.setdefault(
            canonical,
            {"canonical_work": canonical, "source_rows": []},
        )
        target["source_rows"].append(source_row)
    return [grouped[name] for name in SUPPORTED_WORKS if name in grouped]


def _prepare_llm_insights(
    api: Any,
    root: Path,
    manifest: dict[str, Any],
    visual_artifact: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    estimate = read_json(output_path(root, "estimate_normalized.json"), code="estimate_required")
    geometry = read_json(output_path(root, "geometry.json"), code="geometry_required")
    mapping = read_json(output_path(root, "mapping.json"), code="mapping_required")
    quantities = read_json(output_path(root, "quantities.json"), code="audit_result_required")
    trace = read_json(output_path(root, "calculation_trace.json"), code="audit_result_required")
    price_catalog = read_json(output_path(root, "price_catalog.json"), code="price_catalog_required")
    findings_artifact = read_json(output_path(root, "findings.json"), code="audit_result_required")
    price_artifact = read_json(output_path(root, "price_checks.json"), code="audit_result_required")
    context_package: dict[str, Any] | None = None
    llm_delegation: dict[str, Any] | None = None
    result: dict[str, Any] | None = None
    # Сокращаем только transport-выборку для аналитика. Полные coverage и summary
    # уже посчитаны Python и остаются в контексте независимо от размера выборки.
    for findings_limit in range(insights.MAX_CONTEXT_FINDINGS, 0, -1):
        candidate_context = insights.build_context_package(
            manifest=manifest,
            estimate=estimate,
            geometry=geometry,
            mapping=mapping,
            quantities=quantities,
            trace=trace,
            findings=findings_artifact["findings"],
            warnings=findings_artifact["warnings"],
            checked_rows=findings_artifact["checked_rows"],
            not_checked_rows=findings_artifact["not_checked_rows"],
            price_checks=price_artifact["checks"],
            price_catalog=price_catalog,
            visual_insights=visual_artifact,
            max_context_findings=findings_limit,
        )
        try:
            candidate_delegation = insights.delegation(candidate_context)
        except ValueError:
            continue
        candidate_result = {
            "status": "llm_insights_required",
            "llm_insights_delegation": candidate_delegation,
            "summary": findings_artifact.get("summary", {}),
            "visual_summary": {
                "status": visual_artifact.get("status"),
                "photos_count": visual_artifact.get("photos_count", 0),
                "insights_count": len(visual_artifact.get("items", [])),
            },
            "llm_context_sha256": canonical_json_sha256(candidate_context),
            "next_action": "schedule_llm_insights_subagent",
            "assistant_instruction": LLM_INSIGHTS_ASSISTANT_INSTRUCTION,
        }
        if len(canonical_json_text(candidate_result)) <= 14_500:
            context_package = candidate_context
            llm_delegation = candidate_delegation
            result = candidate_result
            break
    if context_package is None or llm_delegation is None or result is None:
        raise AuditError(
            "llm_context_too_large",
            "Пакет аналитического этапа превышает безопасный transport-размер.",
        )
    context_sha256 = canonical_json_sha256(context_package)
    atomic_write_json(output_path(root, "llm_context.json"), context_package)
    manifest["audit_status"] = "llm_insights_required"
    manifest["llm_context_sha256"] = context_sha256
    manifest["llm_insights_attempts"] = 0
    manifest["visual_insights_sha256"] = canonical_json_sha256(visual_artifact)
    manifest["visual_review_completed_at"] = now_iso()
    save_manifest(root, manifest)
    result["llm_context_sha256"] = context_sha256
    return manifest, result


def skip_visual_review(api: Any, job_id: str) -> tuple[dict[str, Any], dict[str, Any]]:
    root = job_dir(api, job_id)
    manifest = load_manifest(root)
    if manifest.get("audit_status") != "visual_review_required":
        raise AuditError(
            "visual_review_not_pending",
            "Пропуск фотографий допустим только после детерминированного run_audit.",
        )
    artifact = visual.empty_artifact()
    atomic_write_json(output_path(root, "visual_insights.json"), artifact)
    manifest["visual_review_status"] = "skipped"
    return _prepare_llm_insights(api, root, manifest, artifact)


def import_site_photos(
    api: Any,
    job_id: str,
    archive_descriptor: Any,
) -> tuple[dict[str, Any], dict[str, Any]]:
    root = job_dir(api, job_id)
    manifest = load_manifest(root)
    if manifest.get("audit_status") != "visual_review_required":
        raise AuditError(
            "visual_review_not_pending",
            "ZIP с фотографиями принимается только после детерминированного run_audit.",
        )
    archive_path, archive_name, _ = _validate_attachment_descriptor(archive_descriptor, {".zip"})
    archive_hash = sha256_file(archive_path)
    try:
        with zipfile.ZipFile(archive_path, "r") as archive:
            entries = [
                entry for entry in archive.infolist()
                if not entry.is_dir()
                and not Path(entry.filename).name.startswith(".")
                and Path(entry.filename).suffix.lower() in {".png", ".jpg", ".jpeg"}
            ]
            if not 1 <= len(entries) <= visual.MAX_PHOTOS:
                raise AuditError(
                    "invalid_photo_archive",
                    f"ZIP должен содержать от 1 до {visual.MAX_PHOTOS} фотографий PNG/JPG/JPEG.",
                    {"photos_found": len(entries)},
                )
            photo_dir = root / "assets" / "site_photos"
            photo_dir.mkdir(parents=True, exist_ok=True)
            photos: list[dict[str, Any]] = []
            for index, entry in enumerate(entries, start=1):
                suffix = Path(entry.filename).suffix.lower()
                payload = archive.read(entry)
                # Имя из ZIP сохраняем только как метаданные. На диск пишем под
                # контролируемым именем, поэтому ../ в entry не задаёт путь назначения.
                destination = photo_dir / f"photo_{index:03d}{suffix}"
                atomic_write_bytes(destination, payload)
                mime = _image_mime(destination, suffix)
                source_hash = sha256_file(destination)
                projection = _site_photo_projection(api, destination, suffix, source_hash)
                if projection.parent.name != "construction_audit_mvp" or projection.parent.parent.name != "uploads":
                    raise AuditError(
                        "vision_projection_unavailable",
                        "Не удалось создать разрешённый upload projection для фото Vision-субагента.",
                    )
                photos.append({
                    "photo_id": f"photo_{index:03d}",
                    "filename": Path(entry.filename).name,
                    "sha256": source_hash,
                    "size_bytes": len(payload),
                    "mime": mime,
                    "imported_name": destination.name,
                    "vision_source_path": str(projection),
                    "delegation_token": secrets.token_urlsafe(32),
                })
    except zipfile.BadZipFile as exc:
        raise AuditError("invalid_photo_archive", "Attachment не является корректным ZIP-архивом.") from exc

    estimate = read_json(output_path(root, "estimate_normalized.json"), code="estimate_required")
    mapping = read_json(output_path(root, "mapping.json"), code="mapping_required")
    estimate_works = _visual_estimate_works(estimate, mapping)
    visual_delegations = [
        {"photo_id": photo["photo_id"], "packet": visual.delegation(photo, estimate_works)}
        for photo in photos
    ]
    response = {
        "status": "visual_analysis_required",
        "photos_count": len(photos),
        "visual_delegations": visual_delegations,
        "next_action": "schedule_visual_subagents_in_parallel",
        "assistant_instruction": VISUAL_DELEGATION_ASSISTANT_INSTRUCTION,
    }
    response_size = len(canonical_json_text(response))
    if response_size > MAX_VISUAL_DELEGATION_RESPONSE_CHARS:
        raise AuditError(
            "visual_delegation_transport_too_large",
            "Пакеты фотоанализа превышают безопасный transport-размер.",
            {"response_chars": response_size, "maximum": MAX_VISUAL_DELEGATION_RESPONSE_CHARS},
        )
    manifest.setdefault("documents", {})["site_photos_archive"] = {
        "filename": archive_name,
        "sha256": archive_hash,
        "size_bytes": archive_path.stat().st_size,
        "photos_count": len(photos),
    }
    manifest["audit_status"] = "visual_analysis_required"
    manifest["visual_review_status"] = "analysis_required"
    manifest["visual_photos_count"] = len(photos)
    manifest["visual_rejected_task_ids"] = {}
    atomic_write_json(output_path(root, "visual_photos.json"), {
        "schema_version": visual.SCHEMA_VERSION,
        "archive_sha256": archive_hash,
        "estimate_works": estimate_works,
        "photos": photos,
    })
    atomic_write_json(output_path(root, "visual_photo_analyses.json"), {
        "schema_version": visual.SCHEMA_VERSION,
        "items": [],
    })
    save_manifest(root, manifest)
    return manifest, response


def save_visual_analysis(
    api: Any,
    ctx: Any,
    job_id: str,
    photo_id: Any,
    photo_task_id: Any,
    analysis_payload: Any,
) -> tuple[dict[str, Any], dict[str, Any]]:
    root = job_dir(api, job_id)
    manifest = load_manifest(root)
    if manifest.get("audit_status") != "visual_analysis_required":
        raise AuditError("visual_analysis_not_pending", "Сейчас результаты фото Vision не ожидаются.")
    if not isinstance(photo_task_id, str) or not SUBAGENT_TASK_ID_RE.fullmatch(photo_task_id.strip()):
        raise AuditError(
            "visual_analysis_invalid",
            "photo_task_id должен быть точным 8-символьным hex task_id из schedule_subagent.",
        )
    normalized_task_id = photo_task_id.strip()
    parent_task_id = _ctx_value(ctx, "task_id")
    if parent_task_id and normalized_task_id == parent_task_id:
        raise AuditError(
            "visual_subagent_required",
            "Фотоанализ разрешено сохранять только из отдельного Vision-субагента.",
        )
    analysis_object = _decode_subagent_json_object(
        analysis_payload,
        code="visual_analysis_transport_invalid",
        label="analysis",
    )
    photos_artifact = read_json(output_path(root, "visual_photos.json"), code="visual_photos_required")
    photos = photos_artifact.get("photos", [])
    photo = next((item for item in photos if item.get("photo_id") == photo_id), None)
    if photo is None:
        raise AuditError("unknown_photo_id", "photo_id отсутствует в импортированном ZIP.")
    analyses_artifact = read_json(
        output_path(root, "visual_photo_analyses.json"), code="visual_photos_required"
    )
    analyses = analyses_artifact.get("items", [])
    if any(item.get("photo_id") == photo_id for item in analyses):
        raise AuditError("duplicate_visual_analysis", "Результат для этой фотографии уже сохранён.")
    if any(item.get("photo_task_id") == normalized_task_id for item in analyses):
        raise AuditError(
            "duplicate_visual_task_id",
            "Один photo_task_id нельзя использовать для результатов разных фотографий.",
        )
    estimate_works = photos_artifact.get("estimate_works", [])
    rejected_by_photo = manifest.get("visual_rejected_task_ids")
    if not isinstance(rejected_by_photo, dict):
        rejected_by_photo = {}
    rejected_task_ids = rejected_by_photo.get(str(photo_id))
    if not isinstance(rejected_task_ids, list):
        rejected_task_ids = []
    if normalized_task_id in rejected_task_ids:
        # После schema error основной агент не должен «починить» JSON и выдать его
        # за прежний ответ Vision. Тот же отклонённый task_id повторно не принимаем.
        raise AuditError(
            "visual_subagent_retry_required",
            "После schema validation error нужен новый запуск фото Vision-субагента с новым task ID.",
            {
                "photo_id": photo_id,
                "rejected_photo_task_id": normalized_task_id,
                "allowed_next_action": "rerun_visual_subagent",
                "next_action": "rerun_visual_subagent",
                "assistant_instruction": VISUAL_RETRY_ASSISTANT_INSTRUCTION,
                "visual_delegation": visual.delegation(photo, estimate_works),
            },
        )
    try:
        validated = visual.validate(
            analysis_object,
            photo=photo,
            estimate_works=estimate_works,
        )
    except visual.VisualValidationError as exc:
        if normalized_task_id not in rejected_task_ids:
            rejected_task_ids.append(normalized_task_id)
        rejected_by_photo[str(photo_id)] = rejected_task_ids
        manifest["visual_rejected_task_ids"] = rejected_by_photo
        save_manifest(root, manifest)
        raise AuditError(
            "visual_analysis_schema_invalid",
            "Результат фото Vision-субагента не прошёл validation.",
            {
                "photo_id": photo_id,
                "validation_errors": exc.errors,
                "allowed_next_action": "rerun_visual_subagent",
                "next_action": "rerun_visual_subagent",
                "assistant_instruction": VISUAL_RETRY_ASSISTANT_INSTRUCTION,
                "visual_delegation": visual.delegation(
                    photo, estimate_works, validation_errors=exc.errors
                ),
            },
        ) from exc
    analyses.append({
        **validated,
        "photo_task_id": normalized_task_id,
        "photo_sha256": photo["sha256"],
        "saved_at": now_iso(),
    })
    atomic_write_json(output_path(root, "visual_photo_analyses.json"), {
        "schema_version": visual.SCHEMA_VERSION,
        "items": analyses,
    })
    remaining = len(photos) - len(analyses)
    if remaining:
        # Валидные фото сохраняем по одному: ошибка следующего субагента не должна
        # заставить повторять уже принятые результаты.
        save_manifest(root, manifest)
        return manifest, {
            "status": "visual_analysis_required",
            "saved_photo_id": photo_id,
            "remaining_count": remaining,
            "next_action": "save_remaining_visual_analyses",
        }
    artifact = visual.aggregate(photos, analyses)
    atomic_write_json(output_path(root, "visual_insights.json"), artifact)
    manifest["visual_review_status"] = "completed"
    return _prepare_llm_insights(api, root, manifest, artifact)


def finalize_audit(
    api: Any,
    ctx: Any,
    job_id: str,
    insights_task_id: Any,
    insights_payload: Any,
) -> tuple[dict[str, Any], dict[str, Any]]:
    root = job_dir(api, job_id)
    manifest = load_manifest(root)
    if manifest.get("audit_status") != "llm_insights_required":
        raise AuditError(
            "llm_insights_not_pending",
            "Сначала выполните детерминированный run_audit и обязательный LLM-этап.",
        )
    if not isinstance(insights_task_id, str) or not insights_task_id.strip() or len(insights_task_id.strip()) > 200:
        raise AuditError("llm_insights_execution_failed", "insights_task_id должен быть непустой строкой до 200 символов.")
    normalized_task_id = insights_task_id.strip()
    parent_task_id = _ctx_value(ctx, "task_id")
    if parent_task_id and normalized_task_id == parent_task_id:
        raise AuditError(
            "llm_insights_subagent_required",
            "Аналитические наблюдения разрешено сохранять только из отдельного субагента.",
        )
    insights_object = _decode_subagent_json_object(
        insights_payload,
        code="llm_insights_transport_invalid",
        label="llm_insights",
    )

    context_package = read_json(output_path(root, "llm_context.json"), code="llm_context_required")
    # Пока аналитик работал, downstream-артефакты могли быть пересобраны. Ответ
    # принимаем только для того context package, который был ему делегирован.
    if (
        not isinstance(context_package, dict)
        or canonical_json_sha256(context_package) != manifest.get("llm_context_sha256")
    ):
        raise AuditError("llm_context_changed", "Контекст аналитического этапа изменился после расчётов.")
    attempts = int(manifest.get("llm_insights_attempts") or 0)
    if attempts >= MAX_LLM_INSIGHTS_ATTEMPTS:
        raise AuditError(
            "llm_insights_analysis_failed",
            "Аналитический этап не прошёл validation после трёх попыток.",
            {"attempts": attempts, "maximum_attempts": MAX_LLM_INSIGHTS_ATTEMPTS},
        )
    try:
        validated_insights = insights.validate(insights_object, context_package)
    except insights.InsightsValidationError as exc:
        attempts += 1
        manifest["llm_insights_attempts"] = attempts
        save_manifest(root, manifest)
        if attempts >= MAX_LLM_INSIGHTS_ATTEMPTS:
            raise AuditError(
                "llm_insights_analysis_failed",
                "Аналитический этап не прошёл validation после трёх попыток.",
                {
                    "validation_errors": exc.errors,
                    "allowed_next_action": "stop",
                    "attempts": attempts,
                    "maximum_attempts": MAX_LLM_INSIGHTS_ATTEMPTS,
                },
            ) from exc
        retry_delegation = insights.delegation(
            context_package,
            validation_errors=exc.errors,
        )
        raise AuditError(
            "llm_insights_schema_invalid",
            "Результат аналитического субагента не прошёл validation.",
            {
                "validation_errors": exc.errors,
                "allowed_next_action": "rerun_llm_insights_subagent",
                "attempts": attempts,
                "maximum_attempts": MAX_LLM_INSIGHTS_ATTEMPTS,
                "llm_insights_delegation": retry_delegation,
                "assistant_instruction": (
                    "Вызови schedule_subagent с llm_insights_delegation из этого ответа строго 1:1, "
                    "затем wait_task. Не исправляй JSON аналитического субагента самостоятельно."
                ),
            },
        ) from exc

    manifest["llm_insights_attempts"] = attempts + 1
    completed_at = now_iso()
    insights_artifact = {
        **validated_insights,
        "context_sha256": manifest["llm_context_sha256"],
        "insights_task_id": normalized_task_id,
        "generated_at": completed_at,
    }
    atomic_write_json(output_path(root, "llm_insights.json"), insights_artifact)

    estimate = read_json(output_path(root, "estimate_normalized.json"), code="estimate_required")
    geometry = read_json(output_path(root, "geometry.json"), code="geometry_required")
    mapping = read_json(output_path(root, "mapping.json"), code="mapping_required")
    quantities = read_json(output_path(root, "quantities.json"), code="audit_result_required")
    trace = read_json(output_path(root, "calculation_trace.json"), code="audit_result_required")
    price_catalog = read_json(output_path(root, "price_catalog.json"), code="price_catalog_required")
    findings_artifact = read_json(output_path(root, "findings.json"), code="audit_result_required")
    price_artifact = read_json(output_path(root, "price_checks.json"), code="audit_result_required")
    visual_artifact = read_json(output_path(root, "visual_insights.json"), code="visual_review_required")
    # Фото входят в аналитический контекст отдельной выборкой, поэтому их hash
    # сверяем ещё раз непосредственно перед сборкой HTML.
    if canonical_json_sha256(visual_artifact) != manifest.get("visual_insights_sha256"):
        raise AuditError("visual_insights_changed", "Visual insights изменились после подготовки LLM-контекста.")
    findings = findings_artifact["findings"]
    warnings = findings_artifact["warnings"]
    checked_rows = findings_artifact["checked_rows"]
    not_checked_rows = findings_artifact["not_checked_rows"]
    price_checks = price_artifact["checks"]
    quantity_checks = findings_artifact.get("quantity_checks") or []
    summary = findings_artifact.get("summary") or audit_summary(
        estimate, findings, warnings, checked_rows, not_checked_rows, price_checks,
        quantity_checks,
    )

    manifest["audit_completed"] = True
    manifest["report_generated"] = True
    manifest["audit_status"] = "completed"
    manifest["audit_completed_at"] = completed_at
    manifest["llm_insights_task_id"] = normalized_task_id
    manifest["llm_insights_sha256"] = canonical_json_sha256(insights_artifact)
    manifest["llm_insights_status"] = validated_insights["status"]
    manifest["llm_insights_completed_at"] = completed_at
    manifest["skill_version"] = SKILL_VERSION

    html = report.build_report(
        {
            "manifest": manifest,
            "estimate": estimate,
            "geometry": geometry,
            "mapping": mapping,
            "quantities": quantities,
            "calculation_trace": trace,
            "findings": findings,
            "warnings": warnings,
            "checked_rows": checked_rows,
            "not_checked_rows": not_checked_rows,
            "quantity_checks": quantity_checks,
            "price_catalog": price_catalog,
            "price_checks": price_checks,
            "summary": summary,
            "llm_insights": insights_artifact,
            "visual_insights": visual_artifact,
        }
    )
    report_path = output_path(root, "report.html")
    atomic_write_bytes(report_path, html.encode("utf-8"))
    save_manifest(root, manifest)
    audit_summary_markdown = build_audit_summary_markdown(
        summary, findings, warnings, geometry, report_path, visual_artifact, insights_artifact
    )
    return manifest, {
        "status": "audit_completed",
        "audit_summary_markdown": audit_summary_markdown,
        "summary": summary,
        "report_artifact": {
            "path": str(report_path),
            "size_bytes": report_path.stat().st_size,
            "sha256": sha256_file(report_path),
        },
        "next_action": "display_audit_summary_markdown_verbatim",
        "assistant_instruction": AUDIT_SUMMARY_ASSISTANT_INSTRUCTION,
    }
